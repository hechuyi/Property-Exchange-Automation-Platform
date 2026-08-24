from __future__ import annotations

import hashlib
import json
import os
import tempfile
import unittest
from dataclasses import fields
from datetime import datetime as _REAL_DATETIME
from unittest.mock import patch

from openpyxl import load_workbook
from path_isolation import assert_peap_env_under_temp, isolated_peap_env

from desktop_backend.services.records_service import RecordsService
from peap.export_evidence_policy import export_evidence_verdict_accepted
from peap.export_projection import ExportProjectionError, append_export_projection_findings
from peap.streaming_export import (
    IMPUTED_DEAL_DATE_REMARK_SUFFIX,
    _artifact_checksums,
    _default_cursor_id,
    _field_missing_diagnostic,
    _normalized_business_scope_ids,
    _project_deal_rows,
    _projection_error_missing_fields,
    _record_business_id,
    _record_source_id,
    _write_value_row,
    count_records_in_export_scope_by_state,
    record_to_export_payload,
    run_ready_export,
)
from peap.streaming_ingest import StreamingIngestRunner
from peap.streaming_models import (
    ExportArtifact,
    ExportRequest,
    ExportRunResult,
    IngestedRecord,
    ItemSavedPayload,
    PostProcessFinding,
)
from peap.streaming_store import StreamingStore


class StreamingExportWriterBoundaryTest(unittest.TestCase):
    def test_artifact_checksums_fail_closed_for_missing_and_symlink_files(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            missing = os.path.join(temp_dir, "missing.xlsx")
            missing_artifact = ExportArtifact("equity_transfer", "new", missing, 1)
            with self.assertRaisesRegex(ExportProjectionError, "missing"):
                _artifact_checksums([missing_artifact])

            source = os.path.join(temp_dir, "source.xlsx")
            link = os.path.join(temp_dir, "link.xlsx")
            with open(source, "wb") as handle:
                handle.write(b"xlsx")
            try:
                os.symlink(source, link)
            except (AttributeError, NotImplementedError, OSError):
                self.skipTest("symbolic links are unavailable")
            symlink_artifact = ExportArtifact("equity_transfer", "new", link, 1)
            with self.assertRaisesRegex(ExportProjectionError, "regular file"):
                _artifact_checksums([symlink_artifact])

    def test_write_value_row_rejects_non_mapping_rows(self) -> None:
        for row in ([], False, None):
            with self.subTest(row=row):
                with self.assertRaisesRegex(TypeError, "row"):
                    _write_value_row(row, kind="equity_transfer")

    def test_write_value_row_keeps_export_placeholders_blank(self) -> None:
        row = {
            "项目编号": "G32026SH1000999",
            "项目名称": "占位符清洗项目",
            "类型": "--",
            "隶属集团": "--",
            "转让方": "--",
            "挂牌价格": "--",
            "挂牌开始日期": "2026-06-01",
        }

        values = _write_value_row(row, kind="equity_transfer")
        headers = [
            "类型",
            "项目编号",
            "隶属集团",
            "转让方",
            "项目名称",
            "挂牌价格",
            "所属行业",
            "挂牌开始日期",
            "挂牌截止日期",
            "受托机构",
            "交易所",
            "经办人",
            "近一年净利润（万）",
            "所在地区",
            "挂牌次数",
            "备注",
        ]

        exported = dict(zip(headers, values, strict=True))
        self.assertEqual(exported["类型"], "")
        self.assertEqual(exported["隶属集团"], "")
        self.assertEqual(exported["转让方"], "")
        self.assertEqual(exported["挂牌价格"], "")
        self.assertEqual(exported["挂牌开始日期"], "2026-06-01")

    def test_project_deal_rows_rejects_explicit_non_mapping_export_extras(self) -> None:
        payload = {"项目编号": "G62026BJ000001", "项目名称": "测试增资项目", "成交日期": "2026-04-18"}
        for export_extras in ([], False):
            with self.subTest(export_extras=export_extras):
                record = {
                    "canonical_record": {
                        "source_identity": {"source_id": "cbex"},
                        "canonical_fields": {},
                        "export_extras": export_extras,
                    },
                }

                with self.assertRaisesRegex(TypeError, "export_extras"):
                    _project_deal_rows(
                        record,
                        payload,
                        business_id="deal_capital_increase",
                        output_kind="deal_capital_increase",
                    )

    def test_append_export_projection_findings_rejects_explicit_non_mapping_finding_evidence(self) -> None:
        with self.assertRaisesRegex(TypeError, "evidence"):
            append_export_projection_findings(
                [
                    PostProcessFinding(
                        severity="warn",
                        type="canonical_field_missing",
                        message="bad evidence shape",
                        evidence=[],  # type: ignore[arg-type]
                    )
                ],
                {
                    "record_family": "listing",
                    "canonical_fields": {
                        "project_code": "",
                    },
                },
            )

    def test_append_export_projection_findings_rejects_scalar_missing_fields_evidence(self) -> None:
        with self.assertRaisesRegex(TypeError, r"missing_fields"):
            append_export_projection_findings(
                [
                    PostProcessFinding(
                        severity="warn",
                        type="canonical_field_missing",
                        message="bad missing fields shape",
                        evidence={"missing_fields": "project_code"},
                    )
                ],
                {
                    "record_family": "listing",
                    "canonical_fields": {
                        "project_code": "",
                    },
                },
            )

    def test_field_missing_diagnostic_rejects_scalar_record_findings(self) -> None:
        record = {
            "record_id": "rec-bad-findings",
            "revision_id": 7,
            "record_family": "listing",
            "business_id": "equity_transfer",
            "project_code": "G32026SH1000998",
            "project_name": "坏发现列表项目",
            "findings": "bad findings",
        }

        with self.assertRaisesRegex(TypeError, "findings"):
            _field_missing_diagnostic(record)

    def test_field_missing_diagnostic_rejects_non_mapping_record_finding(self) -> None:
        record = {
            "record_id": "rec-bad-finding-item",
            "revision_id": 7,
            "record_family": "listing",
            "business_id": "equity_transfer",
            "project_code": "G32026SH1000997",
            "project_name": "坏发现项项目",
            "findings": ["not a finding object"],
        }

        with self.assertRaisesRegex(TypeError, r"record\.findings\[\*\]"):
            _field_missing_diagnostic(record)

    def test_field_missing_diagnostic_rejects_scalar_missing_fields_argument(self) -> None:
        record = {
            "record_id": "rec-bad-missing-fields-argument",
            "revision_id": 7,
            "record_family": "listing",
            "business_id": "equity_transfer",
        }

        with self.assertRaisesRegex(TypeError, "missing_fields"):
            _field_missing_diagnostic(record, missing_fields="类型")  # type: ignore[arg-type]

    def test_field_missing_diagnostic_preserves_canonical_failure_code(self) -> None:
        diagnostic = _field_missing_diagnostic(
            {
                "record_id": "rec-canonical-missing",
                "revision_id": 7,
                "record_family": "listing",
                "business_id": "equity_transfer",
                "findings": [
                    {
                        "type": "canonical_field_missing",
                        "evidence": {"missing_fields": ["price"]},
                    }
                ],
            }
        )

        self.assertEqual(diagnostic["failure_code"], "canonical_field_missing")

    def test_projection_error_missing_fields_accepts_none(self) -> None:
        exc = ExportProjectionError("projection failed")

        self.assertEqual(_projection_error_missing_fields(exc), [])

    def test_projection_error_missing_fields_rejects_scalar_missing_fields(self) -> None:
        exc = ExportProjectionError("projection failed")
        object.__setattr__(exc, "missing_fields", "类型")

        with self.assertRaisesRegex(TypeError, "missing_fields"):
            _projection_error_missing_fields(exc)

    def test_normalized_business_scope_ids_rejects_scalar_business_types(self) -> None:
        with self.assertRaisesRegex(TypeError, "business_types"):
            _normalized_business_scope_ids("股权转让", record_family="listing")  # type: ignore[arg-type]

    def test_normalized_business_scope_ids_rejects_mapping_business_types(self) -> None:
        with self.assertRaisesRegex(TypeError, "business_types"):
            _normalized_business_scope_ids({"type": "股权转让"}, record_family="listing")  # type: ignore[arg-type]


class StreamingExportTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.env_patch = patch.dict(os.environ, isolated_peap_env(self.temp_dir.name))
        self.env_patch.start()
        self.addCleanup(self.env_patch.stop)
        assert_peap_env_under_temp(self, self.temp_dir.name)
        self.store = StreamingStore(f"{self.temp_dir.name}/streaming.sqlite3", auto_migrate=True)
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-1",
                revision_hash="hash-1",
                project_code="G32025SH1000194",
                project_name="测试项目",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-21",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/a.html",
                archive_path=self._artifact_path("a.html"),
                parser_payload={
                    "项目编号": "G32025SH1000194",
                    "项目名称": "测试项目",
                    "项目类型": "股权转让",
                    "类型": "国资",
                    "转让方": "上海测试公司",
                    "挂牌次数": 3,
                    "挂牌开始日期": "2026-03-21",
                },
                postprocess_payload={"项目编号": "G32025SH1000194", "项目名称": "测试项目", "项目类型": "股权转让"},
                canonical_record={
                    "record_family": "listing",
                    "canonical_fields": {
                        "project_code": "G32025SH1000194",
                        "project_name": "测试项目",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "shanghai",
                        "start_date": "2026-03-21",
                        "price": "108.00",
                        "seller": "上海测试公司",
                        "source_type": "国资",
                        "group_name": "上海电气集团",
                        "listing_times": 3,
                    },
                },
                canonical_projection={
                    "项目编号": "G32025SH1000194",
                    "项目名称": "测试项目",
                    "项目类型": "股权转让",
                    "项目状态": "挂牌中",
                    "类型": "国资",
                    "转让方": "上海测试公司",
                    "隶属集团": "上海电气集团",
                    "挂牌开始日期": "2026-03-21",
                    "挂牌价格": "108.00",
                    "挂牌次数": 3,
                },
                findings=[],
            )
        )

    def _artifact_path(self, name: str, content: bytes = b"fixture artifact") -> str:
        path = os.path.join(self.temp_dir.name, "archive", name)
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "wb") as handle:
            handle.write(content)
        return path

    def _write_test_export_artifact(self, file_path: str) -> None:
        from openpyxl import Workbook

        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        workbook = Workbook()
        workbook.active.append(["ok"])
        workbook.save(file_path)

    def _write_evidence_sidecar(self, artifact_path: str, payload: dict[str, object]) -> None:
        with open(f"{artifact_path}.peap-evidence.json", "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, sort_keys=True)

    @staticmethod
    def _exchange_label(source_id: str) -> str:
        return {
            "cbex": "北交所",
            "sse": "上交所",
            "tpre": "天交所",
            "cquae": "重交所",
        }[source_id]

    def _upsert_deal_record(
        self,
        *,
        record_id: str,
        revision_hash: str,
        source_id: str,
        business_id: str,
        project_code: str,
        project_name: str,
        project_type: str,
        deal_date: str = "2026-04-18",
        collection_date: str | None = None,
        listing_date: str | None = None,
        deal_date_basis: str = "deal_date",
        deal_date_is_imputed: bool = False,
        deal_price: str = "1080.5",
        valuation: str = "1200",
        reserve_price: str = "1000",
        export_extras: dict[str, object] | None = None,
    ) -> None:
        canonical_extras = dict(export_extras or {})
        resolved_collection_date = deal_date if collection_date is None else collection_date
        resolved_listing_date = deal_date if listing_date is None else listing_date
        self.store.upsert_record(
            IngestedRecord(
                record_id=record_id,
                revision_hash=revision_hash,
                project_code=project_code,
                project_name=project_name,
                project_type=project_type,
                exchange=source_id,
                listing_date=resolved_listing_date,
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/{record_id}.html",
                archive_path=self._artifact_path(f"{record_id}.html"),
                parser_payload={"project_code": project_code, "project_name": project_name},
                postprocess_payload={"project_code": project_code, "project_name": project_name},
                canonical_record={
                    "record_family": "deal",
                    "source_identity": {"source_id": source_id},
                    "business_identity": {
                        "project_code": project_code,
                        "business_id": business_id,
                        "raw_business_label": project_type,
                    },
                    "canonical_fields": {
                        "project_code": project_code,
                        "project_name": project_name,
                        "project_type": project_type,
                        "status": "成交",
                        "exchange": self._exchange_label(source_id),
                        "deal_date": deal_date,
                        "deal_date_basis": deal_date_basis,
                        "deal_date_is_imputed": deal_date_is_imputed,
                        "collection_date": resolved_collection_date,
                        "deal_price": deal_price,
                        "deal_price_unit_basis": "default_wan",
                        "valuation": valuation,
                        "reserve_price": reserve_price,
                    },
                    "export_extras": canonical_extras,
                },
                findings=[],
                record_family="deal",
                source_identity={
                    "record_family": "deal",
                    "source_id": source_id,
                    "business_id": business_id,
                    "project_code": project_code,
                    "project_name": project_name,
                    "listing_date": resolved_listing_date,
                },
            )
        )

    def test_export_request_contract_exposes_requested_mode_and_cursor_id_only(self) -> None:
        field_names = {item.name for item in fields(ExportRequest)}

        self.assertIn("requested_export_mode", field_names)
        self.assertIn("cursor_id", field_names)
        self.assertNotIn("mode", field_names)
        self.assertNotIn("cursor_key", field_names)

        result_field_names = {item.name for item in fields(ExportRunResult)}
        self.assertIn("field_missing_diagnostics", result_field_names)
        self.assertNotIn("skipped_incomplete", result_field_names)
        self.assertNotIn("incomplete_diagnostics", result_field_names)

    def test_run_ready_export_uses_unique_export_id_within_same_second(self) -> None:
        request = ExportRequest(
            date_from="2026-03-01",
            date_to="2026-03-31",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        class _FrozenDatetime(_REAL_DATETIME):
            @classmethod
            def now(cls, *_args, **_kwargs):
                return _REAL_DATETIME(2026, 3, 22, 10, 0, 0)

        writer_calls: list[str] = []

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            writer_calls.append(file_path)

        with patch("peap.streaming_export.dt.datetime", _FrozenDatetime):
            first = run_ready_export(self.store, request, writer=fake_writer)
            second = run_ready_export(self.store, request, writer=fake_writer)

        self.assertNotEqual(first.export_id, second.export_id)
        self.assertEqual(len(writer_calls), 1)

    def test_record_business_id_does_not_fall_back_to_project_type(self) -> None:
        self.assertEqual(
            _record_business_id(
                {
                    "record_family": "listing",
                    "project_type": "股权转让",
                    "canonical_record": {
                        "record_family": "listing",
                        "business_identity": {},
                    },
                }
            ),
            "",
        )

    def test_record_business_id_uses_canonical_source_identity_after_business_identity(self) -> None:
        self.assertEqual(
            _record_business_id(
                {
                    "record_family": "listing",
                    "canonical_record": {
                        "record_family": "listing",
                        "business_identity": {},
                        "source_identity": {"business_id": "equity_transfer"},
                    },
                }
            ),
            "equity_transfer",
        )

    def test_record_source_id_rejects_explicit_non_mapping_legacy_source_identity(self) -> None:
        for source_identity in ([], False):
            with self.subTest(source_identity=source_identity):
                with self.assertRaisesRegex(TypeError, "source_identity"):
                    _record_source_id(
                        {
                            "source_identity_json": None,
                            "source_identity": source_identity,
                        }
                    )

    def test_field_missing_diagnostic_rejects_scalar_missing_fields_evidence(self) -> None:
        record = {
            "record_id": "rec-bad-missing-fields",
            "revision_id": 7,
            "record_family": "listing",
            "business_id": "equity_transfer",
            "project_code": "G32026SH1000999",
            "project_name": "坏契约项目",
            "findings": [
                {
                    "type": "export_field_missing",
                    "evidence": {"missing_fields": "类型"},
                }
            ],
        }

        with self.assertRaisesRegex(TypeError, r"evidence\.missing_fields"):
            _field_missing_diagnostic(record)

    def test_iter_records_in_export_scope_excludes_field_missing_even_if_acknowledged(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-field-missing",
                revision_hash="hash-field-missing",
                project_code="G32026SH1009999",
                project_name="缺字段项目",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-21",
                state="field_missing",
                source_file=f"{self.temp_dir.name}/raw/missing.html",
                archive_path=self._artifact_path("missing.html"),
                parser_payload={"项目编号": "G32026SH1009999"},
                postprocess_payload={"项目编号": "G32026SH1009999"},
                canonical_record={
                    "record_family": "listing",
                    "canonical_fields": {
                        "project_code": "G32026SH1009999",
                        "project_name": "缺字段项目",
                    },
                },
                findings=[
                    PostProcessFinding(
                        severity="warn",
                        type="export_field_missing",
                        message="导出字段缺失：类型",
                        evidence={},
                    )
                ],
            )
        )

        with self.store._connect() as conn:
            conn.execute(
                "UPDATE records SET acknowledged_payload_json = ? WHERE record_id = ?",
                (
                    json.dumps(
                        {
                            "field_missing": {
                                "acknowledged": True,
                                "missing_fields_hash": "hash-ack",
                            }
                        },
                        ensure_ascii=False,
                    ),
                    "rec-field-missing",
                ),
            )

        request = ExportRequest(
            date_from="2026-03-01",
            date_to="2026-03-31",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        result = run_ready_export(self.store, request)

        self.assertEqual(result.new_records + result.changed_records, 1)
        exported_map = self.store.get_exported_revision_map(result.cursor_id)
        self.assertIn("rec-1", exported_map)
        self.assertNotIn("rec-field-missing", exported_map)

    def test_field_missing_only_scope_counts_as_blocked_not_exported(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-field-missing-only",
                revision_hash="hash-field-missing-only",
                project_code="G32026SH1009998",
                project_name="缺字段项目",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-22",
                state="field_missing",
                source_file=f"{self.temp_dir.name}/raw/missing-only.html",
                archive_path=self._artifact_path("missing-only.html"),
                parser_payload={"项目编号": "G32026SH1009998"},
                postprocess_payload={"项目编号": "G32026SH1009998"},
                canonical_record={
                    "record_family": "listing",
                    "business_identity": {
                        "business_id": "equity_transfer",
                        "raw_business_label": "股权转让",
                    },
                    "canonical_fields": {
                        "project_code": "G32026SH1009998",
                        "project_name": "缺字段项目",
                        "project_type": "股权转让",
                    },
                },
                findings=[
                    PostProcessFinding(
                        severity="warn",
                        type="export_field_missing",
                        message="导出字段缺失：类型",
                        evidence={"missing_fields": ["类型"]},
                    )
                ],
                source_identity={
                    "record_family": "listing",
                    "source_id": "sse",
                    "business_id": "equity_transfer",
                },
            )
        )
        with self.store._connect() as conn:
            conn.execute(
                "UPDATE records SET acknowledged_payload_json = ? WHERE record_id = ?",
                (
                    json.dumps(
                        {
                            "field_missing": {
                                "acknowledged": True,
                                "missing_fields_hash": "hash-ack",
                            }
                        },
                        ensure_ascii=False,
                    ),
                    "rec-field-missing-only",
                ),
            )
        request = ExportRequest(
            date_from="2026-03-22",
            date_to="2026-03-22",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        result = run_ready_export(self.store, request)
        scope_counts = count_records_in_export_scope_by_state(self.store, request)

        self.assertEqual(result.new_records + result.changed_records, 0)
        self.assertEqual(result.field_missing_blocked_records, 1)
        self.assertEqual(scope_counts, {"field_missing": 1})
        diagnostic = result.field_missing_diagnostics[0]
        self.assertEqual(diagnostic["record_id"], "rec-field-missing-only")
        self.assertEqual(diagnostic["revision_id"], 2)
        self.assertEqual(diagnostic["record_family"], "listing")
        self.assertEqual(diagnostic["business_id"], "equity_transfer")
        self.assertEqual(diagnostic["failure_code"], "export_field_missing")
        self.assertEqual(
            diagnostic["missing_fields"],
            [
                {
                    "kind": "export",
                    "field": "类型",
                    "canonical_field": "",
                    "export_field": "类型",
                    "message": "export field 类型 is required",
                }
            ],
        )
        self.assertEqual(self.store.list_exports(limit=10), [])
        self.assertEqual(self.store.get_exported_revision_map(result.cursor_id), {})

    def test_export_demotes_legacy_ready_projection_failure_and_preserves_ack(self) -> None:
        record_id = "rec-legacy-ready-missing-price"
        self.store.upsert_record(
            IngestedRecord(
                record_id=record_id,
                revision_hash="hash-legacy-ready-missing-price",
                project_code="G32026GD1000888",
                project_name="遗留挂牌价格缺失项目",
                project_type="股权转让",
                exchange="guangdong",
                listing_date="2026-08-18",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/legacy-ready-missing-price.html",
                archive_path=self._artifact_path("legacy-ready-missing-price.html"),
                parser_payload={"项目编号": "G32026GD1000888"},
                postprocess_payload={"项目编号": "G32026GD1000888"},
                canonical_record={
                    "record_family": "listing",
                    "business_identity": {
                        "business_id": "equity_transfer",
                        "raw_business_label": "股权转让",
                    },
                    "canonical_fields": {
                        "project_code": "G32026GD1000888",
                        "project_name": "遗留挂牌价格缺失项目",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "guangdong",
                        "start_date": "2026-08-18",
                        "price": "",
                        "seller": "广东测试公司",
                    },
                },
                findings=[],
                record_family="listing",
                source_identity={
                    "record_family": "listing",
                    "source_id": "guangdong",
                    "business_id": "equity_transfer",
                },
            )
        )
        acknowledgement = {
            "field_missing": {
                "acknowledged": True,
                "missing_fields_hash": "pre-existing-ack",
            }
        }
        with self.store._connect() as conn:
            conn.execute(
                "UPDATE records SET acknowledged_payload_json = ? WHERE record_id = ?",
                (json.dumps(acknowledgement, ensure_ascii=False), record_id),
            )
        stored_before_export = self.store.get_record(record_id)
        self.store.mark_mapping_pending(
            record_id=record_id,
            revision_id=int(stored_before_export["revision_id"]),
            project_code="G32026GD1000888",
            payload={"项目编号": "G32026GD1000888"},
        )

        request = ExportRequest(
            date_from="2026-08-18",
            date_to="2026-08-18",
            business_types=["equity_transfer"],
            exchange="guangdong",
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        result = run_ready_export(self.store, request)

        self.assertEqual(result.artifacts, [])
        self.assertEqual(result.field_missing_blocked_records, 1)
        self.assertEqual(result.field_missing_diagnostics[0]["failure_code"], "canonical_field_missing")
        stored = self.store.get_record(record_id)
        self.assertEqual(stored["state"], "field_missing")
        self.assertEqual(stored["acknowledged_payload_json"], acknowledgement)
        self.assertEqual(self.store.list_pending_mappings(), [])
        self.assertEqual(
            [item["type"] for item in stored["findings"]],
            ["canonical_field_missing"],
        )
        self.assertEqual(stored["findings"][0]["evidence"]["missing_fields"], ["price"])
        self.assertIn(
            "canonical_field_missing",
            stored["canonical_record"]["policy_state"]["findings"],
        )
        self.assertEqual(
            count_records_in_export_scope_by_state(self.store, request),
            {"field_missing": 1},
        )

    def test_ready_record_with_corrupt_business_identity_is_blocked_not_hidden(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-ready-corrupt-business-identity",
                revision_hash="hash-ready-corrupt-business-identity",
                project_code="G32026SH1999003",
                project_name="业务身份损坏记录",
                project_type="股权转让",
                exchange="sse",
                listing_date="2026-04-22",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/ready-corrupt-business-identity.html",
                archive_path=self._artifact_path("ready-corrupt-business-identity.html"),
                parser_payload={"项目编号": "G32026SH1999003", "项目名称": "业务身份损坏记录"},
                postprocess_payload={"项目编号": "G32026SH1999003", "项目名称": "业务身份损坏记录"},
                canonical_record={
                    "record_family": "listing",
                    "business_identity": {
                        "business_id": "equity_transfer",
                        "raw_business_label": "股权转让",
                    },
                    "canonical_fields": {
                        "project_code": "G32026SH1999003",
                        "project_name": "业务身份损坏记录",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "sse",
                        "start_date": "2026-04-22",
                        "price": "108.00",
                        "seller": "上海测试公司",
                    },
                },
                findings=[],
                record_family="listing",
                source_identity={
                    "record_family": "listing",
                    "source_id": "sse",
                    "business_id": "equity_transfer",
                },
            )
        )
        corrupted_canonical_record = {
            "record_family": "listing",
            "business_identity": "oops",
            "canonical_fields": {
                "project_code": "G32026SH1999003",
                "project_name": "业务身份损坏记录",
                "project_type": "股权转让",
                "status": "挂牌中",
                "exchange": "sse",
                "start_date": "2026-04-22",
                "price": "108.00",
                "seller": "上海测试公司",
            },
        }
        with self.store._connect() as conn:
            row = conn.execute(
                "SELECT latest_revision_id FROM records WHERE record_id = ?",
                ("rec-ready-corrupt-business-identity",),
            ).fetchone()
            conn.execute(
                "UPDATE records SET business_id = '' WHERE record_id = ?",
                ("rec-ready-corrupt-business-identity",),
            )
            conn.execute(
                "UPDATE record_revisions SET canonical_record_json = ? WHERE revision_id = ?",
                (
                    json.dumps(corrupted_canonical_record, ensure_ascii=False, sort_keys=True),
                    int(row["latest_revision_id"]),
                ),
            )
        request = ExportRequest(
            date_from="2026-04-22",
            date_to="2026-04-22",
            business_types=None,
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        result = run_ready_export(self.store, request)
        scope_counts = count_records_in_export_scope_by_state(self.store, request)

        self.assertEqual(result.new_records + result.changed_records, 0)
        self.assertEqual(len(result.artifacts), 0)
        self.assertEqual(scope_counts, {"ready": 1})
        self.assertEqual(result.field_missing_blocked_records, 1)
        self.assertEqual(result.field_missing_diagnostics[0]["record_id"], "rec-ready-corrupt-business-identity")
        self.assertEqual(result.field_missing_diagnostics[0]["failure_code"], "invalid_identity_shape")
        self.assertEqual(self.store.list_exports(limit=10), [])
        self.assertEqual(self.store.get_exported_revision_map(result.cursor_id), {})

    def test_ready_record_with_corrupt_business_identity_and_keyword_is_blocked_not_raised(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-ready-corrupt-business-identity-keyword",
                revision_hash="hash-ready-corrupt-business-identity-keyword",
                project_code="G32026SH1999004",
                project_name="关键字业务身份损坏记录",
                project_type="股权转让",
                exchange="sse",
                listing_date="2026-04-23",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/ready-corrupt-business-identity-keyword.html",
                archive_path=self._artifact_path("ready-corrupt-business-identity-keyword.html"),
                parser_payload={"项目编号": "G32026SH1999004", "项目名称": "关键字业务身份损坏记录"},
                postprocess_payload={"项目编号": "G32026SH1999004", "项目名称": "关键字业务身份损坏记录"},
                canonical_record={
                    "record_family": "listing",
                    "business_identity": {
                        "business_id": "equity_transfer",
                        "raw_business_label": "股权转让",
                    },
                    "canonical_fields": {
                        "project_code": "G32026SH1999004",
                        "project_name": "关键字业务身份损坏记录",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "sse",
                        "start_date": "2026-04-23",
                        "price": "108.00",
                        "seller": "上海测试公司",
                    },
                },
                findings=[],
                record_family="listing",
                source_identity={
                    "record_family": "listing",
                    "source_id": "sse",
                    "business_id": "equity_transfer",
                },
            )
        )
        corrupted_canonical_record = {
            "record_family": "listing",
            "business_identity": "oops",
            "canonical_fields": {
                "project_code": "G32026SH1999004",
                "project_name": "关键字业务身份损坏记录",
                "project_type": "股权转让",
                "status": "挂牌中",
                "exchange": "sse",
                "start_date": "2026-04-23",
                "price": "108.00",
                "seller": "上海测试公司",
            },
        }
        with self.store._connect() as conn:
            row = conn.execute(
                "SELECT latest_revision_id FROM records WHERE record_id = ?",
                ("rec-ready-corrupt-business-identity-keyword",),
            ).fetchone()
            conn.execute(
                "UPDATE records SET business_id = '' WHERE record_id = ?",
                ("rec-ready-corrupt-business-identity-keyword",),
            )
            conn.execute(
                "UPDATE record_revisions SET canonical_record_json = ? WHERE revision_id = ?",
                (
                    json.dumps(corrupted_canonical_record, ensure_ascii=False, sort_keys=True),
                    int(row["latest_revision_id"]),
                ),
            )
        request = ExportRequest(
            date_from="2026-04-23",
            date_to="2026-04-23",
            business_types=None,
            keyword="关键字业务身份损坏记录",
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        result = run_ready_export(self.store, request)

        self.assertEqual(result.new_records + result.changed_records, 0)
        self.assertEqual(len(result.artifacts), 0)
        self.assertEqual(result.field_missing_blocked_records, 1)
        self.assertEqual(
            result.field_missing_diagnostics[0]["record_id"],
            "rec-ready-corrupt-business-identity-keyword",
        )
        self.assertEqual(result.field_missing_diagnostics[0]["failure_code"], "invalid_identity_shape")

    def test_ready_export_requires_verified_artifact_evidence(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-stale-artifact",
                revision_hash="hash-stale-artifact",
                project_code="G32026SH1000770",
                project_name="证据缺失项目",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-23",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/stale-artifact.html",
                archive_path=f"{self.temp_dir.name}/archive/stale-artifact.html",
                parser_payload={"项目编号": "G32026SH1000770", "项目名称": "证据缺失项目"},
                postprocess_payload={"项目编号": "G32026SH1000770", "项目名称": "证据缺失项目"},
                canonical_record={
                    "record_family": "listing",
                    "business_identity": {
                        "business_id": "equity_transfer",
                        "raw_business_label": "股权转让",
                    },
                    "canonical_fields": {
                        "project_code": "G32026SH1000770",
                        "project_name": "证据缺失项目",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "shanghai",
                        "start_date": "2026-03-23",
                        "price": "108.00",
                        "seller": "上海测试公司",
                    },
                },
                findings=[],
                source_identity={
                    "record_family": "listing",
                    "source_id": "sse",
                    "business_id": "equity_transfer",
                },
            )
        )
        request = ExportRequest(
            date_from="2026-03-23",
            date_to="2026-03-23",
            business_types=["equity_transfer"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        result = run_ready_export(self.store, request)

        self.assertEqual(result.new_records + result.changed_records, 0)
        self.assertEqual(result.artifacts, [])
        self.assertEqual(result.field_missing_blocked_records, 1)
        self.assertEqual(
            result.field_missing_diagnostics[0]["failure_code"],
            "artifact_evidence_policy_rejected",
        )
        self.assertEqual(self.store.get_exported_revision_map(result.cursor_id), {})

    def test_ready_export_allows_verified_artifact_evidence(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        result = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))

        self.assertEqual(result.new_records, 1)
        self.assertEqual(result.field_missing_diagnostics, [])
        self.assertEqual(len(result.artifacts), 1)

    def test_export_evidence_policy_acceptance_matrix_is_verdict_only(self) -> None:
        accepted_verdicts = [
            {"status": "verified", "safe_evidence": {}},
            {
                "status": "shared_official_page",
                "safe_evidence": {"page_kind": "shared_official_page"},
            },
            {
                "status": "verified",
                "safe_evidence": {},
                "field_missing_acknowledgement": {"acknowledged": True},
            },
        ]
        rejected_verdicts = [
            {"status": "shared_official_page", "safe_evidence": {"page_kind": "detail_page"}},
            {"status": "shared_official_page", "safe_evidence": {}},
            {"status": "present_unverified", "safe_evidence": {}},
            {"status": "stale_reference", "safe_evidence": {}},
            {"status": "invalid_shell", "safe_evidence": {}},
            {"status": "identity_mismatch", "safe_evidence": {}},
        ]

        for verdict in accepted_verdicts:
            self.assertTrue(export_evidence_verdict_accepted(verdict), verdict)
        for verdict in rejected_verdicts:
            self.assertFalse(export_evidence_verdict_accepted(verdict), verdict)

    def test_export_evidence_policy_rejects_malformed_shared_official_page_safe_evidence(self) -> None:
        with self.assertRaisesRegex(TypeError, "safe_evidence must be a mapping"):
            export_evidence_verdict_accepted(
                {"status": "shared_official_page", "safe_evidence": "shared_official_page"}
            )

    def test_ready_export_allows_explicit_shared_official_page_evidence(self) -> None:
        artifact_path = self.store.get_record("rec-1")["archive_path"]
        with open(artifact_path, "rb") as handle:
            content_hash = "sha256:" + hashlib.sha256(handle.read()).hexdigest()
        self._write_evidence_sidecar(
            artifact_path,
            {
                "schema_version": 1,
                "page_kind": "shared_official_page",
                "content_sha256": content_hash,
                "identity_hints": {
                    "record_family": "listing",
                    "business_id": "equity_transfer",
                    "source_id": "shanghai",
                    "project_code": "G32025SH1000194",
                },
                "source_locator_hash": "sha256:1111",
                "final_locator_hash": "sha256:2222",
            },
        )

        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        result = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))

        self.assertEqual(result.new_records, 1)
        self.assertEqual(result.field_missing_diagnostics, [])

    def test_ready_export_rejects_unaccepted_shared_official_page_metadata(self) -> None:
        artifact_path = self.store.get_record("rec-1")["archive_path"]
        self._write_evidence_sidecar(
            artifact_path,
            {
                "schema_version": 1,
                "page_kind": "shared_official_page",
                "content_sha256": "sha256:not-the-artifact",
                "identity_hints": {
                    "record_family": "listing",
                    "business_id": "equity_transfer",
                    "source_id": "shanghai",
                    "project_code": "G32025SH1000194",
                },
                "source_locator_hash": "sha256:1111",
                "final_locator_hash": "sha256:2222",
            },
        )
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        result = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))

        self.assertEqual(result.new_records + result.changed_records, 0)
        self.assertEqual(result.artifacts, [])
        self.assertEqual(result.field_missing_diagnostics[0]["failure_code"], "artifact_evidence_policy_rejected")
        self.assertEqual(result.field_missing_diagnostics[0]["evidence_status"], "present_unverified")

    def test_ready_export_rejects_invalid_shell_sidecar_without_body_marker(self) -> None:
        artifact_path = self.store.get_record("rec-1")["archive_path"]
        with open(artifact_path, "rb") as handle:
            content_hash = "sha256:" + hashlib.sha256(handle.read()).hexdigest()
        self._write_evidence_sidecar(
            artifact_path,
            {
                "schema_version": 1,
                "page_kind": "invalid_shell",
                "content_sha256": content_hash,
                "identity_hints": {
                    "record_family": "listing",
                    "business_id": "equity_transfer",
                    "source_id": "shanghai",
                    "project_code": "G32025SH1000194",
                },
                "source_locator_hash": "sha256:1111",
                "final_locator_hash": "sha256:2222",
            },
        )
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        result = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))

        self.assertEqual(result.new_records + result.changed_records, 0)
        self.assertEqual(result.artifacts, [])
        self.assertEqual(result.field_missing_diagnostics[0]["failure_code"], "artifact_evidence_policy_rejected")
        self.assertEqual(result.field_missing_diagnostics[0]["evidence_status"], "invalid_shell")

    def test_ready_export_rejects_invalid_shell_artifact_evidence(self) -> None:
        shell_marker = bytes((60, 104, 49, 62, 83, 83, 69, 32, 68, 101, 97, 108, 32, 78, 111, 116, 105, 99, 101, 60, 47, 104, 49, 62))
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-invalid-shell",
                revision_hash="hash-invalid-shell",
                project_code="G32026SH1000771",
                project_name="壳页面项目",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-24",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/invalid-shell.html",
                archive_path=self._artifact_path("invalid-shell.html", shell_marker),
                parser_payload={"项目编号": "G32026SH1000771", "项目名称": "壳页面项目"},
                postprocess_payload={"项目编号": "G32026SH1000771", "项目名称": "壳页面项目"},
                canonical_record={
                    "record_family": "listing",
                    "business_identity": {
                        "business_id": "equity_transfer",
                        "raw_business_label": "股权转让",
                    },
                    "canonical_fields": {
                        "project_code": "G32026SH1000771",
                        "project_name": "壳页面项目",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "shanghai",
                        "start_date": "2026-03-24",
                        "price": "108.00",
                        "seller": "上海测试公司",
                    },
                },
                findings=[],
                source_identity={
                    "record_family": "listing",
                    "source_id": "sse",
                    "business_id": "equity_transfer",
                },
            )
        )
        request = ExportRequest(
            date_from="2026-03-24",
            date_to="2026-03-24",
            business_types=["equity_transfer"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        result = run_ready_export(self.store, request)

        self.assertEqual(result.new_records + result.changed_records, 0)
        self.assertEqual(result.artifacts, [])
        self.assertEqual(
            result.field_missing_diagnostics[0]["failure_code"],
            "artifact_evidence_policy_rejected",
        )

    def test_identity_mismatch_ready_journey_has_no_export_and_user_readable_reason(self) -> None:
        with patch.dict(os.environ, isolated_peap_env(self.temp_dir.name)):
            assert_peap_env_under_temp(self, self.temp_dir.name)
            self.store.upsert_record(
                IngestedRecord(
                    record_id="rec-identity-mismatch",
                    revision_hash="hash-identity-mismatch",
                    project_code="G32026SH1000888",
                    project_name="身份不一致项目",
                    project_type="股权转让",
                    exchange="shanghai",
                    listing_date="2026-03-25",
                    state="ready",
                    source_file=f"{self.temp_dir.name}/raw/identity-mismatch.html",
                    archive_path=self._artifact_path("identity-mismatch.html"),
                    parser_payload={"项目编号": "G32026SH1000888", "项目名称": "身份不一致项目"},
                    postprocess_payload={"项目编号": "G32026SH1000888", "项目名称": "身份不一致项目"},
                    canonical_record={
                        "record_family": "listing",
                        "business_identity": {
                            "business_id": "equity_transfer",
                            "raw_business_label": "股权转让",
                        },
                        "canonical_fields": {
                            "project_code": "G32026SH1000888",
                            "project_name": "身份不一致项目",
                            "project_type": "股权转让",
                            "status": "挂牌中",
                            "exchange": "shanghai",
                            "start_date": "2026-03-25",
                            "price": "108.00",
                            "seller": "上海测试公司",
                        },
                    },
                    findings=[],
                    source_identity={
                        "record_family": "listing",
                        "source_id": "sse",
                        "business_id": "equity_transfer",
                        "project_code": "G32026SH1000999",
                    },
                )
            )
            request = ExportRequest(
                date_from="2026-03-25",
                date_to="2026-03-25",
                business_types=["equity_transfer"],
                requested_export_mode="full",
                output_dir=os.environ["PEAP_EXPORT_ROOT"],
            )

            result = run_ready_export(self.store, request)

        self.assertEqual(result.new_records + result.changed_records, 0)
        self.assertEqual(result.artifacts, [])
        self.assertEqual(result.field_missing_blocked_records, 1)
        diagnostic = result.field_missing_diagnostics[0]
        self.assertEqual(diagnostic["failure_code"], "artifact_evidence_policy_rejected")
        self.assertEqual(diagnostic["evidence_status"], "identity_mismatch")
        self.assertEqual(diagnostic["evidence_reason_code"], "project_code_mismatch")
        self.assertIn("项目编号", diagnostic["message"])
        self.assertEqual(self.store.get_exported_revision_map(result.cursor_id), {})

    def test_export_persists_cursor_value_manifest_and_revision_watermark_contract(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        def fake_writer(file_path: str, _rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)

        result = run_ready_export(self.store, request, writer=fake_writer)
        export_row = self.store.get_export(result.export_id)
        summary = export_row["summary"]
        manifest = self.store.get_export_manifest(result.export_id)
        cursor_value = self.store.get_export_cursor_value(result.cursor_id)

        self.assertIn(":profile-listing-equity_transfer:", result.cursor_id)
        self.assertEqual(summary["revision_watermark"], result.revision_watermark)
        self.assertEqual(summary["manifest"]["revision_watermark"], result.revision_watermark)
        self.assertEqual(manifest["revision_watermark"], result.revision_watermark)
        self.assertEqual(manifest["cursor_id"], result.cursor_id)
        self.assertEqual(manifest["export_profile_id"], "listing/equity_transfer")
        self.assertEqual(manifest["included_count"], 1)
        self.assertEqual(manifest["excluded_count"], 0)
        self.assertEqual(manifest["cursor_basis"]["export_id"], result.export_id)
        self.assertIn(result.artifacts[0].file_path, manifest["artifact_checksums"])
        self.assertEqual(cursor_value["last_successful_revision_watermark"], result.revision_watermark)
        self.assertEqual(cursor_value["last_successful_export_id"], result.export_id)
        self.assertEqual(cursor_value["cursor_basis_export_id"], result.export_id)
        self.assertEqual(cursor_value["export_profile_id"], "listing/equity_transfer")
        self.assertEqual(cursor_value["canonical_scope_hash"], manifest["canonical_scope_hash"])
        self.assertTrue(cursor_value["eligible_set_hash"])

    def test_export_does_not_persist_history_when_artifact_disappears_before_mark_exported(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        def fake_writer(file_path: str, _rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)

        from peap import streaming_export as streaming_export_module

        real_checksums = streaming_export_module._artifact_checksums

        def disappear_after_checksum(artifacts: list[ExportArtifact]) -> dict[str, str]:
            checksums = real_checksums(artifacts)
            os.remove(artifacts[0].file_path)
            return checksums

        with patch(
            "peap.streaming_export._artifact_checksums",
            side_effect=disappear_after_checksum,
        ):
            with self.assertRaisesRegex((RuntimeError, ValueError), "artifact"):
                run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(self.store.list_exports(), [])
        self.assertEqual(self.store.get_exported_revision_map("default"), {})

    def test_get_export_surfaces_invalid_summary_json_instead_of_returning_empty_summary(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        result = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))
        with self.store._connect() as conn:
            conn.execute(
                "UPDATE exports SET summary_json = ? WHERE export_id = ?",
                ("{", result.export_id),
            )

        with self.assertRaises(json.JSONDecodeError):
            self.store.get_export(result.export_id)

    def test_list_exports_surfaces_invalid_summary_json_instead_of_returning_empty_summary(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        result = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))
        with self.store._connect() as conn:
            conn.execute(
                "UPDATE exports SET summary_json = ? WHERE export_id = ?",
                ("{", result.export_id),
            )

        with self.assertRaises(json.JSONDecodeError):
            self.store.list_exports(limit=10)

    def test_export_cursor_manifest_and_audit_commit_atomically(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        class _Unserializable:
            pass

        with self.assertRaises(TypeError):
            run_ready_export(
                self.store,
                request,
                writer=lambda path, _rows: self._write_test_export_artifact(path),
                audit_action="manual_export",
                audit_payload={"bad": _Unserializable()},
            )

        self.assertEqual(self.store.list_exports(limit=10), [])
        self.assertEqual(self.store.get_exported_revision_map(_default_cursor_id(request)), {})
        self.assertEqual(self.store.get_export_cursor_value(_default_cursor_id(request)), {})

        with self.store._connect() as conn:
            audit_count = conn.execute(
                "SELECT COUNT(*) FROM audit_log WHERE action = 'manual_export'"
            ).fetchone()[0]
            manifest_count = conn.execute("SELECT COUNT(*) FROM export_manifests").fetchone()[0]
        self.assertEqual(audit_count, 0)
        self.assertEqual(manifest_count, 0)

    def _assert_invalid_run_ready_export_audit_payload_is_rejected(self, audit_payload: object) -> None:
        output_dir = f"{self.temp_dir.name}/exports"
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=output_dir,
        )

        with self.assertRaises((TypeError, ValueError)):
            run_ready_export(
                self.store,
                request,
                writer=lambda path, _rows: self._write_test_export_artifact(path),
                audit_action="manual_export",
                audit_payload=audit_payload,
            )

        self.assertEqual(self.store.list_exports(limit=10), [])
        self.assertEqual(self.store.get_exported_revision_map(_default_cursor_id(request)), {})
        self.assertEqual(self.store.get_export_cursor_value(_default_cursor_id(request)), {})
        with self.store._connect() as conn:
            audit_count = conn.execute(
                "SELECT COUNT(*) FROM audit_log WHERE action = 'manual_export'"
            ).fetchone()[0]
            manifest_count = conn.execute("SELECT COUNT(*) FROM export_manifests").fetchone()[0]
        self.assertEqual(audit_count, 0)
        self.assertEqual(manifest_count, 0)
        self.assertEqual(os.listdir(output_dir), [])

    def test_run_ready_export_rejects_false_audit_payload_before_commit(self) -> None:
        self._assert_invalid_run_ready_export_audit_payload_is_rejected(False)

    def test_run_ready_export_rejects_list_audit_payload_before_commit(self) -> None:
        self._assert_invalid_run_ready_export_audit_payload_is_rejected([])

    def test_custom_writer_must_create_artifact_before_cursor_commit(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        with self.assertRaisesRegex(ExportProjectionError, "artifact"):
            run_ready_export(self.store, request, writer=lambda *_args, **_kwargs: None)

        self.assertEqual(self.store.list_exports(limit=10), [])
        self.assertEqual(self.store.get_exported_revision_map(_default_cursor_id(request)), {})
        self.assertEqual(self.store.get_export_cursor_value(_default_cursor_id(request)), {})
        with self.store._connect() as conn:
            manifest_count = conn.execute("SELECT COUNT(*) FROM export_manifests").fetchone()[0]
        self.assertEqual(manifest_count, 0)

    def test_custom_writer_must_create_valid_xlsx_before_cursor_commit(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        def invalid_writer(file_path: str, _rows: list[dict[str, object]]) -> None:
            with open(file_path, "wb") as handle:
                handle.write(b"not an xlsx workbook")

        with self.assertRaisesRegex(ExportProjectionError, "artifact"):
            run_ready_export(self.store, request, writer=invalid_writer)

        self.assertEqual(self.store.list_exports(limit=10), [])
        self.assertEqual(self.store.get_exported_revision_map(_default_cursor_id(request)), {})
        self.assertEqual(self.store.get_export_cursor_value(_default_cursor_id(request)), {})
        with self.store._connect() as conn:
            manifest_count = conn.execute("SELECT COUNT(*) FROM export_manifests").fetchone()[0]
        self.assertEqual(manifest_count, 0)
        self.assertEqual(os.listdir(request.output_dir), [])

    def test_run_ready_export_removes_written_artifacts_when_later_workbook_fails(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-physical-cleanup",
                revision_hash="hash-physical-cleanup",
                project_code="CBEX-PHYSICAL-CLEANUP",
                project_name="实物资产清理项目",
                project_type="实物资产",
                exchange="cbex",
                listing_date="2026-03-21",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/physical-cleanup.html",
                archive_path=self._artifact_path("physical-cleanup.html"),
                parser_payload={"项目编号": "CBEX-PHYSICAL-CLEANUP", "项目名称": "实物资产清理项目"},
                postprocess_payload={"项目编号": "CBEX-PHYSICAL-CLEANUP", "项目名称": "实物资产清理项目"},
                canonical_record={
                    "record_family": "listing",
                    "source_identity": {"source_id": "cbex"},
                    "business_identity": {
                        "project_code": "CBEX-PHYSICAL-CLEANUP",
                        "business_id": "physical_asset",
                        "raw_business_label": "实物资产",
                    },
                    "canonical_fields": {
                        "project_code": "CBEX-PHYSICAL-CLEANUP",
                        "project_name": "实物资产清理项目",
                        "project_type": "实物资产",
                        "exchange": "北交所",
                        "start_date": "2026-03-21",
                        "price": "208.00",
                        "seller": "北交所测试转让方",
                        "source_type": "国资",
                    },
                },
                findings=[],
                source_identity={
                    "record_family": "listing",
                    "source_id": "cbex",
                    "business_id": "physical_asset",
                },
            )
        )
        output_dir = f"{self.temp_dir.name}/exports"
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让", "实物资产"],
            requested_export_mode="full",
            output_dir=output_dir,
        )
        writer_calls = 0

        def flaky_writer(file_path: str, _rows: list[dict[str, object]]) -> None:
            nonlocal writer_calls
            writer_calls += 1
            self._write_test_export_artifact(file_path)
            if writer_calls == 2:
                raise RuntimeError("simulated writer failure")

        with self.assertRaisesRegex(RuntimeError, "simulated writer failure"):
            run_ready_export(self.store, request, writer=flaky_writer)

        self.assertEqual(self.store.list_exports(limit=10), [])
        self.assertEqual(os.listdir(output_dir), [])

    def test_run_ready_export_removes_written_artifacts_when_store_commit_fails(self) -> None:
        output_dir = f"{self.temp_dir.name}/exports"
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=output_dir,
        )

        class _Unserializable:
            pass

        with self.assertRaises(TypeError):
            run_ready_export(
                self.store,
                request,
                writer=lambda path, _rows: self._write_test_export_artifact(path),
                audit_action="manual_export",
                audit_payload={"bad": _Unserializable()},
            )

        self.assertEqual(self.store.list_exports(limit=10), [])
        self.assertEqual(os.listdir(output_dir), [])

    def test_run_ready_export_matches_records_saved_with_slash_dates(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-2",
                revision_hash="hash-2",
                project_code="G32025SH1000999",
                project_name="斜杠日期项目",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026/03/20",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/b.html",
                archive_path=self._artifact_path("b.html"),
                parser_payload={"项目编号": "G32025SH1000999", "项目名称": "斜杠日期项目"},
                postprocess_payload={"项目编号": "G32025SH1000999", "项目名称": "斜杠日期项目", "项目类型": "股权转让"},
                canonical_record={
                    "record_family": "listing",
                    "canonical_fields": {
                        "project_code": "G32025SH1000999",
                        "project_name": "斜杠日期项目",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "shanghai",
                        "start_date": "2026-03-20",
                        "price": "88.00",
                        "seller": "测试转让方",
                    },
                },
                canonical_projection={
                    "项目编号": "G32025SH1000999",
                    "项目名称": "斜杠日期项目",
                    "项目类型": "股权转让",
                    "项目状态": "挂牌中",
                    "挂牌开始日期": "2026-03-20",
                    "挂牌价格": "88.00",
                    "转让方": "测试转让方",
                },
                findings=[],
            )
        )
        request = ExportRequest(
            date_from="2026-03-20",
            date_to="2026-03-20",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        writer_calls: list[str] = []

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            writer_calls.append(file_path)

        result = run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(result.new_records, 1)
        self.assertEqual(len(result.artifacts), 1)
        self.assertEqual(len(writer_calls), 1)

    def test_run_ready_export_uses_output_contract_headers_and_values(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["file_path"] = file_path
            captured["rows"] = rows

        result = run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(len(result.artifacts), 1)
        self.assertIn("挂牌_股权转让", str(captured["file_path"]))
        self.assertEqual(len(captured["rows"]), 1)
        row = captured["rows"][0]
        self.assertEqual(row["类型"], "国资")
        self.assertEqual(row["转让方"], "上海测试公司")

    def test_run_ready_export_allows_listing_physical_asset_without_status_when_workbook_fields_are_complete(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-physical-ready",
                revision_hash="hash-physical-ready",
                project_code="CBEX-PHYSICAL-001",
                project_name="实物资产就绪项目",
                project_type="实物资产",
                exchange="cbex",
                listing_date="2026-03-21",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/physical-ready.html",
                archive_path=self._artifact_path("physical-ready.html"),
                parser_payload={"项目编号": "CBEX-PHYSICAL-001", "项目名称": "实物资产就绪项目"},
                postprocess_payload={"项目编号": "CBEX-PHYSICAL-001", "项目名称": "实物资产就绪项目"},
                canonical_record={
                    "record_family": "listing",
                    "source_identity": {"source_id": "cbex"},
                    "business_identity": {
                        "project_code": "CBEX-PHYSICAL-001",
                        "business_id": "physical_asset",
                        "raw_business_label": "实物资产",
                    },
                    "canonical_fields": {
                        "project_code": "CBEX-PHYSICAL-001",
                        "project_name": "实物资产就绪项目",
                        "project_type": "实物资产",
                        "exchange": "北交所",
                        "start_date": "2026-03-21",
                        "price": "208.00",
                        "seller": "北交所测试转让方",
                        "source_type": "国资",
                    },
                },
                findings=[],
                source_identity={
                    "record_family": "listing",
                    "source_id": "cbex",
                    "business_id": "physical_asset",
                },
            )
        )
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["实物资产"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["file_path"] = file_path
            captured["rows"] = rows

        result = run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(result.field_missing_diagnostics, [])
        self.assertEqual(result.new_records, 1)
        self.assertEqual(len(result.artifacts), 1)
        self.assertIn("挂牌_实物资产", str(captured["file_path"]))
        row = captured["rows"][0]
        self.assertEqual(row["项目编号"], "CBEX-PHYSICAL-001")
        self.assertEqual(row["类型"], "国资")
        self.assertNotIn("项目状态", row)

    def test_record_to_export_payload_ignores_projection_without_canonical_record(self) -> None:
        payload = record_to_export_payload(
            {
                "project_code": "G32025SH1000194",
                "project_name": "测试项目",
                "project_type": "股权转让",
                "exchange": "shanghai",
                "canonical_projection": {
                    "项目编号": "G32025SH1000194",
                    "项目名称": "测试项目",
                    "项目类型": "股权转让",
                    "挂牌价格": "108.00",
                    "转让方": "上海测试公司",
                },
                "parser_payload": {
                    "未审计透传字段": "should-not-leak",
                },
                "postprocess_payload": {
                    "另一个透传字段": "still-should-not-leak",
                },
            }
        )

        self.assertEqual(payload, {})

    def test_record_to_export_payload_rejects_explicit_non_mapping_canonical_record(self) -> None:
        for canonical_record in ([], False):
            with self.subTest(canonical_record=canonical_record):
                with self.assertRaisesRegex(TypeError, "canonical_record"):
                    record_to_export_payload(
                        {
                            "project_code": "G32025SH1000194",
                            "project_name": "测试项目",
                            "project_type": "股权转让",
                            "exchange": "shanghai",
                            "canonical_record": canonical_record,
                            "canonical_projection": {
                                "项目编号": "G32025SH1000194",
                                "项目名称": "测试项目",
                                "项目类型": "股权转让",
                                "挂牌价格": "108.00",
                                "转让方": "上海测试公司",
                            },
                        }
                    )

    def test_run_ready_export_rejects_projection_only_record_even_when_projection_is_complete(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-projection-only",
                revision_hash="hash-projection-only",
                project_code="G32025SH1000555",
                project_name="projection-only 项目",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-21",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/projection-only.html",
                archive_path=self._artifact_path("projection-only.html"),
                parser_payload={
                    "项目编号": "G32025SH1000555",
                    "项目名称": "projection-only 项目",
                    "项目类型": "股权转让",
                },
                postprocess_payload={
                    "项目编号": "G32025SH1000555",
                    "项目名称": "projection-only 项目",
                    "项目类型": "股权转让",
                },
                canonical_projection={
                    "项目编号": "G32025SH1000555",
                    "项目名称": "projection-only 项目",
                    "项目类型": "股权转让",
                    "项目状态": "挂牌中",
                    "挂牌开始日期": "2026-03-21",
                    "挂牌价格": "108.00",
                    "转让方": "projection-only 卖方",
                },
                findings=[],
            )
        )
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        result = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))
        # rec-projection-only is skipped (no canonical_record), but rec-1 from setUp exports
        self.assertEqual(result.new_records, 1)

    def test_record_to_export_payload_prefers_canonical_fields_over_stale_projection(self) -> None:
        payload = record_to_export_payload(
            {
                "project_code": "G32025SH1000777",
                "project_name": "原始项目",
                "project_type": "股权转让",
                "exchange": "shanghai",
                "canonical_record": {
                    "record_family": "listing",
                    "canonical_fields": {
                        "project_code": "G32025SH1000777",
                        "project_name": "规范化项目",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "shanghai",
                        "start_date": "2026-03-21",
                        "price": "108.00",
                        "seller": "规范化卖方",
                        "source_type": "国资",
                    },
                },
                "canonical_projection": {
                    "项目编号": "G32025SH1000777",
                    "项目名称": "过期项目名",
                    "项目类型": "股权转让",
                    "转让方": "过期卖方",
                    "挂牌价格": "999.99",
                },
                "parser_payload": {
                    "项目名称": "解析层项目名",
                    "转让方": "解析层卖方",
                    "挂牌价格": "666.66",
                },
                "postprocess_payload": {
                    "项目名称": "后处理项目名",
                    "转让方": "后处理卖方",
                    "挂牌价格": "777.77",
                },
            }
        )

        self.assertEqual(payload["项目名称"], "规范化项目")
        self.assertEqual(payload["转让方"], "规范化卖方")
        self.assertEqual(payload["挂牌价格"], "108.00")
        self.assertEqual(payload["类型"], "国资")
        self.assertNotIn("挂牌次数", payload)

    def test_record_to_export_payload_uses_business_identity_family_for_deal_date_semantics(self) -> None:
        payload = record_to_export_payload(
            {
                "project_code": "D32026SH000201",
                "project_name": "身份族成交项目",
                "project_type": "股权转让",
                "exchange": "sse",
                "canonical_record": {
                    "business_identity": {
                        "record_family": "deal",
                        "business_id": "deal_equity_transfer",
                        "raw_business_label": "股权转让",
                    },
                    "source_identity": {"source_id": "sse"},
                    "canonical_fields": {
                        "project_code": "D32026SH000201",
                        "project_name": "身份族成交项目",
                        "project_type": "股权转让",
                        "status": "成交",
                        "exchange": "上交所",
                        "deal_date": "",
                        "deal_date_basis": "collection_date",
                        "deal_date_is_imputed": True,
                        "collection_date": "2026-04-18",
                        "deal_price": "1080.5",
                        "deal_price_unit_basis": "default_wan",
                        "valuation": "1200",
                        "reserve_price": "1000",
                    },
                },
            }
        )

        self.assertEqual(payload.get("成交日期"), "2026-04-18")
        self.assertIn(IMPUTED_DEAL_DATE_REMARK_SUFFIX, str(payload.get("备注") or ""))

    def test_record_to_export_payload_uses_source_identity_family_for_deal_date_semantics(self) -> None:
        payload = record_to_export_payload(
            {
                "project_code": "D32026SH000202",
                "project_name": "来源族成交项目",
                "project_type": "股权转让",
                "exchange": "sse",
                "canonical_record": {
                    "business_identity": {
                        "business_id": "deal_equity_transfer",
                        "raw_business_label": "股权转让",
                    },
                    "source_identity": {"source_id": "sse", "record_family": "deal"},
                    "canonical_fields": {
                        "project_code": "D32026SH000202",
                        "project_name": "来源族成交项目",
                        "project_type": "股权转让",
                        "status": "成交",
                        "exchange": "上交所",
                        "deal_date": "",
                        "deal_date_basis": "collection_date",
                        "deal_date_is_imputed": True,
                        "collection_date": "2026-04-18",
                        "deal_price": "1080.5",
                        "deal_price_unit_basis": "default_wan",
                        "valuation": "1200",
                        "reserve_price": "1000",
                    },
                },
            }
        )

        self.assertEqual(payload.get("成交日期"), "2026-04-18")
        self.assertIn(IMPUTED_DEAL_DATE_REMARK_SUFFIX, str(payload.get("备注") or ""))

    def test_run_ready_export_prefers_persisted_canonical_projection_over_raw_payload_merge(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-canonical-export",
                revision_hash="hash-canonical-export",
                project_code="G32025SH1000998",
                project_name="原始项目名",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-21",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/canonical.html",
                archive_path=self._artifact_path("canonical.html"),
                parser_payload={
                    "项目编号": "G32025SH1000998",
                    "项目名称": "解析层项目名",
                    "项目类型": "股权转让",
                    "转让方": "解析层卖方",
                },
                postprocess_payload={
                    "项目编号": "G32025SH1000998",
                    "项目名称": "后处理项目名",
                    "项目类型": "股权转让",
                    "转让方": "后处理卖方",
                },
                source_identity={
                    "record_family": "listing",
                    "original_source_file": f"{self.temp_dir.name}/raw/canonical.html",
                    "source_url": "https://example.test/detail/export-canonical",
                    "project_code": "G32025SH1000998",
                    "project_name": "原始项目名",
                    "exchange": "shanghai",
                    "listing_date": "2026-03-21",
                    "candidate_tokens": [
                        "project_code:G32025SH1000998",
                        "page_url:https://example.test/detail/export-canonical",
                    ],
                },
                canonical_record={
                    "record_family": "listing",
                    "business_identity": {"project_code": "G32025SH1000998"},
                    "canonical_fields": {
                        "project_code": "G32025SH1000998",
                        "project_name": "规范化项目名",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "start_date": "2026-03-21",
                        "price": "108.00",
                        "seller": "规范化卖方",
                    },
                },
                canonical_projection={
                    "项目编号": "G32025SH1000998",
                    "项目名称": "规范化项目名",
                    "项目类型": "股权转让",
                    "项目状态": "挂牌中",
                    "挂牌开始日期": "2026-03-21",
                    "挂牌价格": "108.00",
                    "转让方": "规范化卖方",
                },
                findings=[],
            )
        )
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["rows"] = rows

        run_ready_export(self.store, request, writer=fake_writer)

        row = next(item for item in captured["rows"] if item["项目编号"] == "G32025SH1000998")
        self.assertEqual(row["项目名称"], "规范化项目名")
        self.assertEqual(row["转让方"], "规范化卖方")

    def test_record_to_export_payload_does_not_pass_through_arbitrary_raw_fields(self) -> None:
        payload = record_to_export_payload(
            {
                "project_code": "G32025SH1000194",
                "project_name": "测试项目",
                "project_type": "股权转让",
                "exchange": "shanghai",
                "canonical_record": {
                    "record_family": "listing",
                    "canonical_fields": {
                        "project_code": "G32025SH1000194",
                        "project_name": "测试项目",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "shanghai",
                        "start_date": "2026-03-21",
                        "price": "108.00",
                        "seller": "上海测试公司",
                        "source_type": "国资",
                    },
                },
                "parser_payload": {
                    "项目编号": "G32025SH1000194",
                    "项目名称": "测试项目",
                    "项目类型": "股权转让",
                    "挂牌价格": "108.00",
                    "未审计透传字段": "should-not-leak",
                },
                "postprocess_payload": {
                    "转让方": "上海测试公司",
                    "另一个透传字段": "still-should-not-leak",
                },
            }
        )

        self.assertEqual(payload["项目编号"], "G32025SH1000194")
        self.assertEqual(payload["挂牌价格"], "108.00")
        self.assertEqual(payload["转让方"], "上海测试公司")
        self.assertNotIn("未审计透传字段", payload)
        self.assertNotIn("另一个透传字段", payload)

    def test_record_to_export_payload_preserves_public_resource_fields_needed_by_writer(self) -> None:
        payload = record_to_export_payload(
            {
                "project_code": "GR20260001",
                "project_name": "成交样例项目",
                "project_type": "股权转让",
                "exchange": "北交所",
                "canonical_record": {
                    "record_family": "listing",
                    "canonical_fields": {
                        "project_code": "GR20260001",
                        "project_name": "成交样例项目",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "北交所",
                        "start_date": "2026/03/01",
                        "price": "108.00",
                        "seller": "样例受让方",
                    },
                    "export_extras": {
                        "交易方式": "网络竞价",
                        "受让方名称": "样例受让方",
                        "转让标的评估值": "88.00",
                        "成交金额": "108.00",
                        "成交日期": "2026/03/01",
                    },
                },
                "parser_payload": {
                    "交易所": "北交所",
                    "项目编号": "GR20260001",
                    "项目名称": "成交样例项目",
                    "交易方式": "网络竞价",
                    "受让方名称": "样例受让方",
                    "转让标的评估值": "88.00",
                    "成交金额": "108.00",
                    "成交日期": "2026/03/01",
                },
                "postprocess_payload": {},
            }
        )

        self.assertEqual(payload["交易方式"], "网络竞价")
        self.assertEqual(payload["受让方名称"], "样例受让方")
        self.assertEqual(payload["转让标的评估值"], "88.00")
        self.assertEqual(payload["成交金额"], "108.00")
        self.assertEqual(payload["成交日期"], "2026/03/01")

    def test_run_ready_export_full_twice_still_exports_full_scoped_range(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-2",
                revision_hash="hash-2",
                project_code="G32025SH1000195",
                project_name="测试项目二",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-21",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/c.html",
                archive_path=self._artifact_path("c.html"),
                parser_payload={
                    "项目编号": "G32025SH1000195",
                    "项目名称": "测试项目二",
                    "项目类型": "股权转让",
                    "挂牌开始日期": "2026-03-21",
                },
                postprocess_payload={
                    "项目编号": "G32025SH1000195",
                    "项目名称": "测试项目二",
                    "项目类型": "股权转让",
                },
                canonical_record={
                    "record_family": "listing",
                    "canonical_fields": {
                        "project_code": "G32025SH1000195",
                        "project_name": "测试项目二",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "shanghai",
                        "start_date": "2026-03-21",
                        "price": "66.00",
                        "seller": "测试卖方二",
                    },
                },
                canonical_projection={
                    "项目编号": "G32025SH1000195",
                    "项目名称": "测试项目二",
                    "项目类型": "股权转让",
                    "项目状态": "挂牌中",
                    "挂牌开始日期": "2026-03-21",
                    "挂牌价格": "66.00",
                    "转让方": "测试卖方二",
                },
                findings=[],
            )
        )
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        first_capture: dict[str, object] = {}
        second_capture: dict[str, object] = {}

        def fake_writer_first(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            first_capture["file_path"] = file_path
            first_capture["rows"] = rows

        def fake_writer_second(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            second_capture["file_path"] = file_path
            second_capture["rows"] = rows

        first = run_ready_export(self.store, request, writer=fake_writer_first)
        second = run_ready_export(self.store, request, writer=fake_writer_second)

        self.assertEqual(first.new_records, 2)
        self.assertEqual(first.changed_records, 0)
        self.assertEqual(len(first.artifacts), 1)
        self.assertEqual(len(first_capture["rows"]), 2)
        self.assertEqual(second.new_records, 2)
        self.assertEqual(second.changed_records, 0)
        self.assertEqual(len(second.artifacts), 1)
        self.assertEqual(len(second_capture["rows"]), 2)

    def test_run_ready_export_accepts_mixed_case_full(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="Full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        first_capture: dict[str, object] = {}
        second_capture: dict[str, object] = {}

        def fake_writer_first(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            first_capture["rows"] = rows

        def fake_writer_second(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            second_capture["rows"] = rows

        first = run_ready_export(self.store, request, writer=fake_writer_first)
        second = run_ready_export(self.store, request, writer=fake_writer_second)

        self.assertEqual(first.new_records, 1)
        self.assertEqual(second.new_records, 1)
        self.assertEqual(len(first_capture["rows"]), 1)
        self.assertEqual(len(second_capture["rows"]), 1)

    def test_run_ready_export_rejects_legacy_rebuild_without_side_effects(self) -> None:
        output_dir = f"{self.temp_dir.name}/legacy-rebuild-exports"
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="ReBuild",
            output_dir=output_dir,
        )

        with self.assertRaisesRegex(ValueError, "requested_export_mode"):
            run_ready_export(self.store, request)

        self.assertFalse(os.path.exists(output_dir))
        self.assertEqual(self.store.list_exports(limit=10), [])

    def test_default_cursor_id_changes_with_keyword_scope(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        object.__setattr__(request, "requested_state", "all")
        object.__setattr__(request, "keyword", "")

        first = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))

        object.__setattr__(request, "keyword", "北交所")
        second = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))

        self.assertNotEqual(first.cursor_id, second.cursor_id)

    def test_default_cursor_id_stable_across_requested_export_mode(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        object.__setattr__(request, "requested_state", "all")
        object.__setattr__(request, "keyword", "")
        first = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))
        object.__setattr__(request, "requested_export_mode", "full")
        second = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))
        self.assertEqual(first.cursor_id, second.cursor_id)

    def test_default_cursor_id_excludes_output_dir_and_includes_contract_identity(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports-a",
        )
        object.__setattr__(request, "requested_state", "all")
        object.__setattr__(request, "keyword", "")
        first = _default_cursor_id(request)
        object.__setattr__(request, "output_dir", f"{self.temp_dir.name}/exports-b")
        second = _default_cursor_id(request)

        self.assertEqual(first, second)
        self.assertTrue(first.startswith("ready_export:"))
        self.assertIn(":schema-v", first)
        self.assertIn(":headers-v", first)

    def test_requested_export_mode_full_does_not_inherit_incremental_cursor(self) -> None:
        incremental = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        full = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        first = run_ready_export(self.store, incremental, writer=lambda path, _rows: self._write_test_export_artifact(path))
        second = run_ready_export(self.store, incremental, writer=lambda path, _rows: self._write_test_export_artifact(path))
        full_result = run_ready_export(self.store, full, writer=lambda path, _rows: self._write_test_export_artifact(path))

        self.assertEqual(first.new_records, 1)
        self.assertEqual(second.new_records, 0)
        self.assertEqual(full_result.new_records, 1)
        self.assertEqual(full_result.cursor_id, first.cursor_id)

    def test_incremental_export_uses_cursor_value_revision_watermark(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        self.store.mark_exported(
            export_id="exp-watermark-seed",
            cursor_id=_default_cursor_id(request),
            requested_export_mode="incremental",
            date_from="2026-03-21",
            date_to="2026-03-21",
            project_type="equity_transfer",
            output_dir=f"{self.temp_dir.name}/exports",
            summary={
                "new_records": 0,
                "changed_records": 0,
                "revision_watermark": 99,
            },
            records=[],
            cursor_value={
                "export_profile_id": "listing/equity_transfer",
                "canonical_scope_hash": "bypass",
                "schema_version": "schema-v1",
                "header_version": "headers-v1",
                "last_successful_revision_watermark": 99,
                "last_successful_export_id": "exp-watermark-seed",
                "cursor_basis_export_id": "exp-watermark-seed",
                "eligible_set_hash": "seed",
            },
        )
        cursor_value = self.store.get_export_cursor_value(_default_cursor_id(request))
        cursor_value["canonical_scope_hash"] = "bypass"
        with self.store._connect() as conn:
            conn.execute(
                "UPDATE export_cursor_values SET value_json = ? WHERE cursor_id = ?",
                (json.dumps(cursor_value, ensure_ascii=False), _default_cursor_id(request)),
            )

        with self.assertRaisesRegex(ExportProjectionError, "cursor basis"):
            run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))

        cursor_value["canonical_scope_hash"] = self.store.get_export("exp-watermark-seed")["summary"]["manifest"].get(
            "canonical_scope_hash",
            "",
        )
        if not cursor_value["canonical_scope_hash"]:
            cursor_value["canonical_scope_hash"] = _default_cursor_id(request).split(":")[2]
        from peap.streaming_export import _canonical_scope_hash

        cursor_value["canonical_scope_hash"] = _canonical_scope_hash(request)
        with self.store._connect() as conn:
            conn.execute(
                "UPDATE export_cursor_values SET value_json = ? WHERE cursor_id = ?",
                (json.dumps(cursor_value, ensure_ascii=False), _default_cursor_id(request)),
            )

        result = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))

        self.assertEqual(result.new_records, 1)
        self.assertEqual(result.changed_records, 0)
        self.assertEqual(len(result.artifacts), 1)

    def test_incremental_export_does_not_watermark_skip_never_exported_record(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        cursor_id = _default_cursor_id(request)
        from peap.streaming_export import _canonical_scope_hash

        self.store.mark_exported(
            export_id="exp-watermark-gap-seed",
            cursor_id=cursor_id,
            requested_export_mode="incremental",
            date_from="2026-03-21",
            date_to="2026-03-21",
            project_type="equity_transfer",
            output_dir=f"{self.temp_dir.name}/exports",
            summary={"new_records": 0, "changed_records": 0, "revision_watermark": 99},
            records=[],
            cursor_value={
                "export_profile_id": "listing/equity_transfer",
                "canonical_scope_hash": _canonical_scope_hash(request),
                "schema_version": "schema-v1",
                "header_version": "headers-v1",
                "last_successful_revision_watermark": 99,
                "last_successful_export_id": "exp-watermark-gap-seed",
                "cursor_basis_export_id": "exp-watermark-gap-seed",
                "eligible_set_hash": "seed",
            },
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["rows"] = rows

        result = run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(result.new_records, 1)
        self.assertEqual(result.changed_records, 0)
        self.assertEqual(len(captured["rows"]), 1)
        self.assertIn("rec-1", self.store.get_exported_revision_map(cursor_id))

    def test_incremental_export_surfaces_invalid_cursor_value_json_instead_of_treating_cursor_as_empty(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        self.store.mark_exported(
            export_id="exp-corrupt-cursor-seed",
            cursor_id=_default_cursor_id(request),
            requested_export_mode="incremental",
            date_from="2026-03-21",
            date_to="2026-03-21",
            project_type="equity_transfer",
            output_dir=f"{self.temp_dir.name}/exports",
            summary={"new_records": 0, "changed_records": 0, "revision_watermark": 99},
            records=[],
            cursor_value={"last_successful_revision_watermark": 99},
        )
        with self.store._connect() as conn:
            conn.execute(
                "UPDATE export_cursor_values SET value_json = ? WHERE cursor_id = ?",
                ("{", _default_cursor_id(request)),
            )

        with self.assertRaises(json.JSONDecodeError):
            run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))

    def test_incremental_export_rejects_cursor_value_missing_basis_provenance(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        first = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))
        cursor_value = self.store.get_export_cursor_value(first.cursor_id)
        cursor_value.pop("eligible_set_hash")
        with self.store._connect() as conn:
            conn.execute(
                "UPDATE export_cursor_values SET value_json = ? WHERE cursor_id = ?",
                (json.dumps(cursor_value, ensure_ascii=False, sort_keys=True), first.cursor_id),
            )
        writer_calls: list[str] = []

        def fake_writer(file_path: str, _rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            writer_calls.append(file_path)

        with self.assertRaisesRegex(ExportProjectionError, "cursor value"):
            run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(writer_calls, [])
        self.assertEqual(self.store.get_export_cursor_value(first.cursor_id), cursor_value)
        self.assertEqual(len(self.store.list_exports(limit=10)), 1)

    def test_incremental_export_rejects_missing_cursor_value_row_when_history_exists(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        first = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))
        with self.store._connect() as conn:
            conn.execute(
                "DELETE FROM export_cursor_values WHERE cursor_id = ?",
                (first.cursor_id,),
            )

        with self.assertRaisesRegex(ExportProjectionError, "cursor value"):
            run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))

    def test_incremental_export_marks_new_revision_changed_when_legacy_hash_collides(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        first = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))
        self.assertEqual(first.new_records, 1)

        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-1",
                revision_hash="hash-1",
                project_code="G32025SH1000194",
                project_name="测试项目",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-21",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/a.html",
                archive_path=self._artifact_path("a.html"),
                parser_payload={
                    "项目编号": "G32025SH1000194",
                    "项目名称": "测试项目修正",
                    "项目类型": "股权转让",
                    "类型": "国资",
                    "转让方": "上海测试公司",
                    "挂牌次数": 3,
                    "挂牌开始日期": "2026-03-21",
                },
                postprocess_payload={"项目编号": "G32025SH1000194", "项目名称": "测试项目", "项目类型": "股权转让"},
                canonical_record={
                    "record_family": "listing",
                    "canonical_fields": {
                        "project_code": "G32025SH1000194",
                        "project_name": "测试项目修正",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "shanghai",
                        "start_date": "2026-03-21",
                        "price": "108.00",
                        "seller": "上海测试公司",
                        "source_type": "国资",
                        "group_name": "上海电气集团",
                        "listing_times": 3,
                    },
                },
                canonical_projection={
                    "项目编号": "G32025SH1000194",
                    "项目名称": "测试项目修正",
                    "项目类型": "股权转让",
                    "项目状态": "挂牌中",
                    "类型": "国资",
                    "转让方": "上海测试公司",
                    "隶属集团": "上海电气集团",
                    "挂牌开始日期": "2026-03-21",
                    "挂牌价格": "108.00",
                    "挂牌次数": 3,
                },
                findings=[],
            )
        )

        second = run_ready_export(self.store, request, writer=lambda path, _rows: self._write_test_export_artifact(path))

        self.assertEqual(second.new_records, 0)
        self.assertEqual(second.changed_records, 1)

    def test_run_ready_export_supports_deal_record_family(self) -> None:
        self._upsert_deal_record(
            record_id="rec-deal-support",
            revision_hash="hash-deal-support",
            source_id="sse",
            business_id="deal_equity_transfer",
            project_code="D32026SH000100",
            project_name="成交导出支持项目",
            project_type="股权转让",
        )
        request = ExportRequest(
            date_from="2026-04-18",
            date_to="2026-04-18",
            business_types=["deal_equity_transfer"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["file_path"] = file_path
            captured["rows"] = rows

        result = run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(result.new_records, 1)
        self.assertEqual(result.changed_records, 0)
        self.assertEqual(len(result.artifacts), 1)
        self.assertEqual(len(captured["rows"]), 1)
        self.assertEqual(captured["rows"][0]["项目编号"], "D32026SH000100")

    def test_run_ready_export_rejects_deal_record_with_unknown_source_and_does_not_mark_cursor(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-deal-unknown-source",
                revision_hash="hash-deal-unknown-source",
                project_code="D32026PR000001",
                project_name="未知来源成交项目",
                project_type="股权转让",
                exchange="public_resource",
                listing_date="2026-04-18",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/rec-deal-unknown-source.html",
                archive_path=self._artifact_path("rec-deal-unknown-source.html"),
                parser_payload={"project_code": "D32026PR000001", "project_name": "未知来源成交项目"},
                postprocess_payload={"project_code": "D32026PR000001", "project_name": "未知来源成交项目"},
                canonical_record={
                    "record_family": "deal",
                    "source_identity": {"source_id": "public_resource"},
                    "business_identity": {
                        "project_code": "D32026PR000001",
                        "business_id": "deal_equity_transfer",
                        "raw_business_label": "股权转让",
                    },
                    "canonical_fields": {
                        "project_code": "D32026PR000001",
                        "project_name": "未知来源成交项目",
                        "project_type": "股权转让",
                        "status": "成交",
                        "exchange": "public_resource",
                        "deal_date": "2026-04-18",
                        "collection_date": "2026-04-18",
                        "deal_price": "1080.5",
                        "deal_price_unit_basis": "default_wan",
                        "valuation": "1200",
                        "reserve_price": "1000",
                    },
                },
                findings=[],
                record_family="deal",
                source_identity={
                    "record_family": "deal",
                    "source_id": "public_resource",
                    "business_id": "deal_equity_transfer",
                    "project_code": "D32026PR000001",
                    "project_name": "未知来源成交项目",
                    "listing_date": "2026-04-18",
                },
            )
        )
        request = ExportRequest(
            date_from="2026-04-18",
            date_to="2026-04-18",
            business_types=["deal_equity_transfer"],
            requested_export_mode="incremental",
            cursor_id="test:deal:unknown-source",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )
        writer_calls: list[str] = []

        def fake_writer(file_path: str, _rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            writer_calls.append(file_path)

        with self.assertRaises(ExportProjectionError):
            run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(writer_calls, [])
        self.assertEqual(self.store.get_exported_revision_map(request.cursor_id), {})

    def test_run_ready_export_blocks_deal_missing_source_identity_instead_of_exchange_fallback(self) -> None:
        self._upsert_deal_record(
            record_id="rec-deal-missing-source-identity",
            revision_hash="hash-deal-missing-source-identity",
            source_id="sse",
            business_id="deal_equity_transfer",
            project_code="D32026SH000101",
            project_name="缺来源身份成交项目",
            project_type="股权转让",
        )
        canonical_record = {
            "record_family": "deal",
            "business_identity": {
                "project_code": "D32026SH000101",
                "business_id": "deal_equity_transfer",
                "raw_business_label": "股权转让",
            },
            "canonical_fields": {
                "project_code": "D32026SH000101",
                "project_name": "缺来源身份成交项目",
                "project_type": "股权转让",
                "status": "成交",
                "exchange": "上交所",
                "deal_date": "2026-04-18",
                "deal_date_basis": "deal_date",
                "deal_date_is_imputed": False,
                "collection_date": "2026-04-18",
                "deal_price": "1080.5",
                "deal_price_unit_basis": "default_wan",
                "valuation": "1200",
                "reserve_price": "1000",
            },
            "export_extras": {},
        }
        with self.store._connect() as conn:
            row = conn.execute(
                "SELECT latest_revision_id FROM records WHERE record_id = ?",
                ("rec-deal-missing-source-identity",),
            ).fetchone()
            conn.execute(
                "UPDATE records SET source_identity_json = '{}' WHERE record_id = ?",
                ("rec-deal-missing-source-identity",),
            )
            conn.execute(
                "UPDATE record_revisions SET canonical_record_json = ? WHERE revision_id = ?",
                (
                    json.dumps(canonical_record, ensure_ascii=False, sort_keys=True),
                    int(row["latest_revision_id"]),
                ),
            )
        request = ExportRequest(
            date_from="2026-04-18",
            date_to="2026-04-18",
            business_types=["deal_equity_transfer"],
            requested_export_mode="incremental",
            cursor_id="test:deal:missing-source-identity",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )
        writer_calls: list[str] = []

        def fake_writer(file_path: str, _rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            writer_calls.append(file_path)

        result = run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(writer_calls, [])
        self.assertEqual(len(result.artifacts), 0)
        self.assertEqual(result.field_missing_blocked_records, 1)
        self.assertEqual(result.field_missing_diagnostics[0]["record_id"], "rec-deal-missing-source-identity")
        self.assertIn("source_identity", result.field_missing_diagnostics[0]["failure_code"])
        self.assertEqual(self.store.get_exported_revision_map(request.cursor_id), {})

    def test_deal_equity_workbook_uses_fixed_four_source_headers(self) -> None:
        for source_id, code in (
            ("cbex", "D32026BJ000001"),
            ("sse", "D32026SH000001"),
            ("tpre", "D32026TJ000001"),
            ("cquae", "D32026CQ000001"),
        ):
            self._upsert_deal_record(
                record_id=f"rec-deal-equity-{source_id}",
                revision_hash=f"hash-deal-equity-{source_id}",
                source_id=source_id,
                business_id="deal_equity_transfer",
                project_code=code,
                project_name=f"{source_id}-equity",
                project_type="股权转让",
                export_extras={
                    "交易方式": "网络竞价",
                    "受让方名称": "受让方示例",
                    "备注": "备注示例",
                    "是否竞价": "是",
                    "是否成交": "是",
                },
            )

        request = ExportRequest(
            date_from="2026-04-18",
            date_to="2026-04-18",
            business_types=["deal_equity_transfer"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )

        result = run_ready_export(self.store, request)
        self.assertEqual(result.new_records, 4)
        self.assertEqual(len(result.artifacts), 1)

        workbook = load_workbook(result.artifacts[0].file_path)
        self.assertEqual(workbook.sheetnames, ["北交所", "上交所", "天交所", "重交所"])
        self.assertEqual(
            [cell.value for cell in workbook["北交所"][1]],
            ["项目编号", "标的名称", "转让标的评估结果", "转让底价（万元）", "交易价格（万元）", "成交日期", "备注", "交易方式", "受让方名称"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["上交所"][1]],
            ["项目编号", "标的名称", "转让标的评估值", "转让底价", "交易价格", "成交日期", "备注", "是否竞价"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["天交所"][1]],
            ["项目编号", "标的名称", "转让标的评估值", "转让底价（万元）", "交易价格（万元）", "成交日期", "备注", "是否成交"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["重交所"][1]],
            ["项目编号", "标的名称", "转让标的评估值", "转让底价（万元）", "交易价格（万元）", "成交日期", "备注"],
        )
        self.assertEqual(workbook["上交所"]["H2"].value, "是")
        self.assertEqual(workbook["天交所"]["H2"].value, "是")

    def test_deal_physical_workbook_uses_asset_sheet_names_and_headers(self) -> None:
        for source_id, code in (
            ("cbex", "GR2026BJ000001"),
            ("sse", "GR2026SH000001"),
        ):
            self._upsert_deal_record(
                record_id=f"rec-deal-physical-{source_id}",
                revision_hash=f"hash-deal-physical-{source_id}",
                source_id=source_id,
                business_id="deal_physical_asset",
                project_code=code,
                project_name=f"{source_id}-physical",
                project_type="实物资产",
            )

        request = ExportRequest(
            date_from="2026-04-18",
            date_to="2026-04-18",
            business_types=["deal_physical_asset"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )

        result = run_ready_export(self.store, request)
        self.assertEqual(result.new_records, 2)
        self.assertEqual(len(result.artifacts), 1)

        workbook = load_workbook(result.artifacts[0].file_path)
        self.assertEqual(
            workbook.sheetnames,
            ["北交所资产成交项目", "上交所资产成交项目"],
        )
        header = ["项目编号", "标的名称", "转让标的评估结果", "转让底价", "交易价格", "成交日期", "备注"]
        self.assertEqual([cell.value for cell in workbook["北交所资产成交项目"][1]], header)
        self.assertEqual([cell.value for cell in workbook["上交所资产成交项目"][1]], header)

    def test_default_ingest_deal_physical_imputed_date_keeps_remark_visible_across_records_and_export(self) -> None:
        snapshot_path = f"{self.temp_dir.name}/raw/deal_physical_sse_imputed.html"
        os.makedirs(os.path.dirname(snapshot_path), exist_ok=True)
        metadata = {
            "record_family": "deal",
            "business_id": "deal_physical_asset",
            "source_id": "sse",
            "source_url": "https://example.invalid/sse/physical-imputed",
            "project_code": "GR2026SH009950",
            "project_name": "上交所资产成交补齐项目",
        }
        detail_payload = {
            "xmbh": "GR2026SH009950",
            "xmmc": "上交所资产成交补齐项目",
            "xmlx": "实物资产",
            "cjjg": "6800",
            "pgjz": "7000",
            "zrdf": "6600",
            "fbsj": "2026-04-20",
            "remark": "原始备注",
        }
        with open(snapshot_path, "w", encoding="utf-8") as handle:
            handle.write(
                "<html><head><meta charset='utf-8'></head><body>"
                "<script id='deal_metadata' type='application/json'>"
                + json.dumps(metadata, ensure_ascii=False)
                + "</script>"
                "<script id='deal_detail' type='application/json'>"
                + json.dumps(detail_payload, ensure_ascii=False)
                + "</script>"
                "</body></html>"
            )

        ingest_runner = StreamingIngestRunner(
            store=self.store,
            archive_root=f"{self.temp_dir.name}/submission",
        )
        ingest_result = ingest_runner.ingest(ItemSavedPayload(source_file=snapshot_path, exchange="sse"))
        self.assertEqual(ingest_result["state"], "ready")

        latest = self.store.get_record(str(ingest_result["record_id"]))
        canonical_fields = dict(latest["canonical_record"]["canonical_fields"])
        self.assertEqual(canonical_fields.get("deal_date"), "")
        self.assertEqual(canonical_fields.get("deal_date_basis"), "collection_date")
        self.assertTrue(bool(canonical_fields.get("deal_date_is_imputed")))
        self.assertEqual(canonical_fields.get("collection_date"), "2026/04/20")
        self.assertEqual(
            latest["canonical_record"]["export_extras"].get("备注"),
            "原始备注；成交日期缺失，按采集日填列",
        )

        records_service = RecordsService(store=self.store, db_path=self.store.db_path)
        with patch(
            "desktop_backend.services.records_service._resolve_record_artifact_path",
            return_value="",
        ):
            records_payload = records_service.list_records(
                {
                    "record_family": "deal",
                    "business_id": "deal_physical_asset",
                    "exchange": "sse",
                    "state": "all",
                }
            )

        self.assertEqual(records_payload["display_columns"][-1], "备注")
        self.assertEqual(records_payload["rows"][0]["display_values"]["成交日期"], "2026/04/20")
        self.assertEqual(
            records_payload["rows"][0]["display_values"]["备注"],
            "原始备注；成交日期缺失，按采集日填列",
        )

        request = ExportRequest(
            date_from="2026-04-20",
            date_to="2026-04-20",
            business_types=["deal_physical_asset"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )
        result = run_ready_export(self.store, request)

        workbook = load_workbook(result.artifacts[0].file_path)
        sheet = workbook["上交所资产成交项目"]
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            ["项目编号", "标的名称", "转让标的评估结果", "转让底价", "交易价格", "成交日期", "备注"],
        )
        self.assertEqual(sheet["F2"].value, "2026/04/20")
        self.assertEqual(sheet["G2"].value, "原始备注；成交日期缺失，按采集日填列")

    def test_deal_physical_workbook_rejects_source_without_current_baseline_sheet(self) -> None:
        self._upsert_deal_record(
            record_id="rec-deal-physical-tpre-unsupported",
            revision_hash="hash-deal-physical-tpre-unsupported",
            source_id="tpre",
            business_id="deal_physical_asset",
            project_code="GR2026TJ000001",
            project_name="天交资产成交项目",
            project_type="实物资产",
        )

        request = ExportRequest(
            date_from="2026-04-18",
            date_to="2026-04-18",
            business_types=["deal_physical_asset"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )

        with self.assertRaises(ExportProjectionError):
            run_ready_export(self.store, request)

    def test_deal_capital_export_expands_investors_and_excludes_summary_rows(self) -> None:
        self._upsert_deal_record(
            record_id="rec-deal-capital-expand",
            revision_hash="hash-deal-capital-expand",
            source_id="sse",
            business_id="deal_capital_increase",
            project_code="G62026SH000001",
            project_name="上交增资成交项目",
            project_type="增资扩股",
            deal_price="5000",
            export_extras={
                "investors": [
                    {"name": "投资方甲", "amount": "3000", "ratio": "60%"},
                    {"name": "总计：1000万元", "amount": "5000"},
                    {"name": "小计", "amount": "5000"},
                    {"name": "subtotal", "amount": "5000"},
                    {"name": "投资方乙", "amount": "2000", "ratio": "40%"},
                ],
                "financing_party_names": ["上交融资方A", "上交融资方B"],
                "capital_increase_company_name": "上交融资方A",
                "transferors": ["转让方甲", "转让方乙"],
                "project_parties": [
                    {"label": "转让方", "name": "转让方甲"},
                    {"label": "融资方", "name": "上交融资方A"},
                ],
                "投资总金额（万元）": "5000",
                "持股占比": "35%",
            },
        )
        request = ExportRequest(
            date_from="2026-04-18",
            date_to="2026-04-18",
            business_types=["deal_capital_increase"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["rows"] = rows

        result = run_ready_export(self.store, request, writer=fake_writer)
        self.assertEqual(result.new_records, 1)
        self.assertEqual(result.artifacts[0].record_count, 1)
        self.assertEqual(len(captured["rows"]), 2)
        rows = list(captured["rows"])
        self.assertEqual([row["投资方名称"] for row in rows], ["投资方甲", "投资方乙"])
        self.assertTrue(all(str(row["投资方名称"]) not in {"总计：1000万元", "小计", "subtotal"} for row in rows))
        self.assertTrue(all(row["投资总金额（万元）"] == "5000" for row in rows))
        self.assertTrue(all(row["持股占比"] == "35%" for row in rows))
        self.assertTrue(all(row["增资企业名称"] == "上交融资方A；上交融资方B" for row in rows))
        self.assertTrue(all(isinstance(row.get("investors"), list) for row in rows))
        self.assertTrue(all(isinstance(row.get("transferors"), list) for row in rows))
        self.assertTrue(all(isinstance(row.get("financing_party_names"), list) for row in rows))
        self.assertTrue(all(isinstance(row.get("project_parties"), list) for row in rows))

    def test_deal_capital_export_rejects_present_non_list_investors(self) -> None:
        record = {
            "canonical_record": {
                "source_identity": {"source_id": "sse"},
                "canonical_fields": {},
                "export_extras": {
                    "investors": {"name": "投资方甲", "amount": "3000"},
                    "投资总金额（万元）": "5000",
                },
            },
        }
        payload = {
            "项目编号": "G62026SH000099",
            "项目名称": "上交增资成交坏投资人结构项目",
            "成交日期": "2026-04-18",
        }

        with self.assertRaisesRegex(ExportProjectionError, r"investors.*list"):
            _project_deal_rows(
                record,
                payload,
                business_id="deal_capital_increase",
                output_kind="deal_capital_increase",
            )

    def test_default_ingest_deal_capital_snapshot_exports_expanded_investors_and_keeps_audit(self) -> None:
        snapshot_path = f"{self.temp_dir.name}/raw/ingest_deal_capital_snapshot.html"
        os.makedirs(os.path.dirname(snapshot_path), exist_ok=True)
        metadata = {
            "record_family": "deal",
            "business_id": "deal_capital_increase",
            "source_id": "sse",
            "source_url": "https://www.suaee.com/si/notice/getNoticeDetail?xmid=XM-CAPITAL-001",
            "project_code": "G62026SH009901",
            "project_name": "默认链路增资成交项目",
            "deal_date": "2026-04-20",
            "deal_date_basis": "collection_date",
            "deal_date_is_imputed": True,
            "deal_date_remark_suffix": "成交日期缺失，按采集日填列",
            "remark_suffix": "成交日期缺失，按采集日填列",
            "collection_date": "2026-04-20",
        }
        detail_payload = {
            "xmbh": "G62026SH009901",
            "xmmc": "默认链路增资成交项目",
            "xmlx": "增资扩股",
            "cjrq": "2026-04-20",
            "fbsj": "2026-04-20",
            "cjjg": "5000",
            "pgjz": "5100",
            "zrdf": "4800",
            "investors": [
                {"name": "投资方甲", "amount": "3000", "ratio": "60%"},
                {"name": "总计", "amount": "5000"},
                {"name": "投资方乙", "amount": "2000", "ratio": "40%"},
            ],
            "financingPartyNames": ["融资方A", "融资方B"],
            "capitalIncreaseCompanyName": "融资方A",
            "projectParties": [
                {"label": "融资方", "name": "融资方A"},
                {"label": "转让方", "name": "转让方甲"},
            ],
            "transferors": ["转让方甲"],
        }
        with open(snapshot_path, "w", encoding="utf-8") as handle:
            handle.write(
                "<html><head><meta charset='utf-8'></head><body>"
                "<h1>SSE Deal Notice</h1>"
                "<script id='deal_metadata' type='application/json'>"
                + json.dumps(metadata, ensure_ascii=False)
                + "</script>"
                "<script id='deal_detail' type='application/json'>"
                + json.dumps(detail_payload, ensure_ascii=False)
                + "</script>"
                "</body></html>"
            )

        ingest_runner = StreamingIngestRunner(
            store=self.store,
            archive_root=f"{self.temp_dir.name}/submission",
        )
        ingest_result = ingest_runner.ingest(ItemSavedPayload(source_file=snapshot_path, exchange="sse"))
        self.assertEqual(ingest_result["state"], "ready")

        latest = self.store.get_record(str(ingest_result["record_id"]))
        canonical_fields = dict(latest["canonical_record"]["canonical_fields"])
        self.assertEqual(canonical_fields.get("deal_date_basis"), "collection_date")
        self.assertTrue(bool(canonical_fields.get("deal_date_is_imputed")))
        self.assertEqual(canonical_fields.get("deal_price"), "5000")

        export_extras = dict(latest["canonical_record"].get("export_extras") or {})
        self.assertTrue(isinstance(export_extras.get("investors"), list))
        self.assertTrue(isinstance(export_extras.get("transferors"), list))
        self.assertTrue(isinstance(export_extras.get("financing_party_names"), list))
        self.assertTrue(isinstance(export_extras.get("project_parties"), list))

        request = ExportRequest(
            date_from="2026-04-20",
            date_to="2026-04-20",
            business_types=["deal_capital_increase"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["rows"] = rows

        with patch("peap.streaming_export.export_evidence_verdict_accepted", return_value=True):
            export_result = run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(export_result.new_records, 1)
        self.assertEqual(len(captured["rows"]), 2)
        rows = list(captured["rows"])
        self.assertEqual([row["投资方名称"] for row in rows], ["投资方甲", "投资方乙"])
        self.assertTrue(all(str(row.get("成交日期") or "").strip() for row in rows))
        self.assertTrue(all(str(row.get("投资金额（万元）") or "").strip() for row in rows))
        self.assertTrue(all(isinstance(row.get("investors"), list) for row in rows))
        self.assertTrue(all(isinstance(row.get("financing_party_names"), list) for row in rows))

    def test_default_ingest_cbex_deal_capital_without_structured_investors_still_exports_rows(self) -> None:
        snapshot_path = f"{self.temp_dir.name}/raw/cbex_deal_capital_top_level_investor.html"
        os.makedirs(os.path.dirname(snapshot_path), exist_ok=True)
        metadata = {
            "record_family": "deal",
            "business_id": "deal_capital_increase",
            "source_id": "cbex",
            "source_url": "https://example.invalid/cbex/deal/capital/001",
            "project_code": "G62026BJ009901",
            "project_name": "北交增资成交主链路项目",
            "deal_date": "2026-04-20",
            "collection_date": "2026-04-20",
        }
        detail_payload = {
            "projectCode": "G62026BJ009901",
            "projectName": "北交增资成交主链路项目",
            "businessType": "增资扩股",
            "dealDate": "2026-04-20",
            "collectionDate": "2026-04-20",
            "dealPrice": "8800",
            "valuation": "9000",
            "reservePrice": "8600",
            "投资方名称": "北京产业投资集团有限公司",
            "投资金额（万元）": 3000,
            "持股比例": 12.5,
            "investors": [
                {"name": "北京产业投资集团有限公司"},
            ],
        }
        with open(snapshot_path, "w", encoding="utf-8") as handle:
            handle.write(
                "<html><head><meta charset='utf-8'></head><body>"
                "<h1>CBEX Deal Capital Snapshot</h1>"
                "<script id='deal_metadata' type='application/json'>"
                + json.dumps(metadata, ensure_ascii=False)
                + "</script>"
                "<script id='deal_detail' type='application/json'>"
                + json.dumps(detail_payload, ensure_ascii=False)
                + "</script>"
                "</body></html>"
            )

        ingest_runner = StreamingIngestRunner(
            store=self.store,
            archive_root=f"{self.temp_dir.name}/submission",
        )
        ingest_result = ingest_runner.ingest(ItemSavedPayload(source_file=snapshot_path, exchange="cbex"))
        self.assertEqual(ingest_result["state"], "ready")

        latest = self.store.get_record(str(ingest_result["record_id"]))
        export_extras = dict(latest["canonical_record"].get("export_extras") or {})
        self.assertEqual(
            export_extras.get("investors"),
            [{"name": "北京产业投资集团有限公司", "amount": "3000", "ratio": "12.5"}],
        )

        request = ExportRequest(
            date_from="2026-04-20",
            date_to="2026-04-20",
            business_types=["deal_capital_increase"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["rows"] = rows

        export_result = run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(export_result.new_records, 1)
        self.assertEqual(len(captured["rows"]), 1)
        row = captured["rows"][0]
        self.assertEqual(row["投资方名称"], "北京产业投资集团有限公司")
        self.assertEqual(row["投资金额（万元）"], "3000")
        self.assertEqual(row["持股比例"], "12.5")

    def test_default_ingest_deal_equity_snapshots_export_source_specific_columns(self) -> None:
        fixtures = (
            (
                "cbex",
                "Q32026BJ009901",
                {
                    "projectCode": "Q32026BJ009901",
                    "projectName": "北交股权成交专属列",
                    "businessType": "股权转让",
                    "dealPrice": "8800",
                    "valuation": "9000",
                    "reservePrice": "8600",
                    "dealDate": "2026-04-20",
                    "collectionDate": "2026-04-20",
                    "dealMethod": "网络竞价",
                    "buyerName": "北交受让方",
                },
                {"交易方式": "网络竞价", "受让方名称": "北交受让方"},
            ),
            (
                "sse",
                "Q32026SH009902",
                {
                    "xmbh": "Q32026SH009902",
                    "xmmc": "上交股权成交专属列",
                    "xmlx": "股权转让",
                    "cjjg": "6800",
                    "pgjz": "7000",
                    "zrdf": "6600",
                    "cjrq": "2026-04-20",
                    "fbsj": "2026-04-20",
                    "isAuction": "是",
                },
                {"是否竞价": "是"},
            ),
            (
                "tpre",
                "Q32026TJ009903",
                {
                    "projectCode": "Q32026TJ009903",
                    "projectName": "天交股权成交专属列",
                    "bizType": "PROPERTY_RIGHT_TRANSFER",
                    "dealAmount": "5200",
                    "valuationValue": "5300",
                    "reservePrice": "5000",
                    "contractSignTime": "2026-04-20",
                    "collectionDate": "2026-04-20",
                    "isDeal": "是",
                },
                {"是否成交": "是"},
            ),
        )
        for source_id, project_code, detail_payload, _expected in fixtures:
            snapshot_path = f"{self.temp_dir.name}/raw/equity_{source_id}.html"
            os.makedirs(os.path.dirname(snapshot_path), exist_ok=True)
            metadata = {
                "record_family": "deal",
                "business_id": "deal_equity_transfer",
                "source_id": source_id,
                "source_url": f"https://example.invalid/{source_id}/{project_code}",
                "project_code": project_code,
                "project_name": str(detail_payload.get("projectName") or detail_payload.get("xmmc") or ""),
            }
            with open(snapshot_path, "w", encoding="utf-8") as handle:
                handle.write(
                    "<html><head><meta charset='utf-8'></head><body>"
                    "<script id='deal_metadata' type='application/json'>"
                    + json.dumps(metadata, ensure_ascii=False)
                    + "</script>"
                    "<script id='deal_detail' type='application/json'>"
                    + json.dumps(detail_payload, ensure_ascii=False)
                    + "</script>"
                    "</body></html>"
                )

            ingest_runner = StreamingIngestRunner(
                store=self.store,
                archive_root=f"{self.temp_dir.name}/submission",
            )
            ingest_result = ingest_runner.ingest(ItemSavedPayload(source_file=snapshot_path, exchange=source_id))
            self.assertEqual(ingest_result["state"], "ready")

        request = ExportRequest(
            date_from="2026-04-20",
            date_to="2026-04-20",
            business_types=["deal_equity_transfer"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["rows"] = rows

        export_result = run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(export_result.new_records, 3)
        rows_by_code = {str(row["项目编号"]): row for row in captured["rows"]}
        for _source_id, project_code, _detail_payload, expected_values in fixtures:
            for field_name, expected_value in expected_values.items():
                self.assertEqual(rows_by_code[project_code].get(field_name), expected_value)

    def test_deal_capital_workbook_uses_source_specific_headers_and_keeps_empty_sheet_headers(self) -> None:
        self._upsert_deal_record(
            record_id="rec-deal-capital-header-cbex",
            revision_hash="hash-deal-capital-header-cbex",
            source_id="cbex",
            business_id="deal_capital_increase",
            project_code="G62026BJ000011",
            project_name="北交增资表头项目",
            project_type="增资扩股",
            export_extras={
                "investors": [
                    {"name": "投资方甲", "amount": "4200", "ratio": "58.3%"},
                    {"name": "投资方乙", "amount": "3000", "ratio": "41.7%"},
                ],
                "投资总金额（万元）": "7200",
                "持股占比（%）": "25%",
                "备注": "项目级备注",
            },
        )
        request = ExportRequest(
            date_from="2026-04-18",
            date_to="2026-04-18",
            business_types=["deal_capital_increase"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )

        result = run_ready_export(self.store, request)
        workbook = load_workbook(result.artifacts[0].file_path)

        self.assertEqual(
            workbook.sheetnames,
            ["北交所增资项目", "上海联交所增资项目", "天交所增资扩股项目成交", "重交所增资扩股项目成交"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["北交所增资项目"][1]],
            ["项目编号", "项目名称", "成交日期", "投资方名称", "投资金额（万元）", "持股比例", "投资总金额（万元）", "持股占比", "备注"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["上海联交所增资项目"][1]],
            ["项目编号", "项目名称", "增资企业名称", "成交日期", "投资方名称", "投资金额（万元）", "持股比例", "投资总金额（万元）", "持股占比", "备注"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["天交所增资扩股项目成交"][1]],
            ["项目编号", "项目名称", "成交日期", "投资方名称", "投资金额（万元）", "持股比例", "投资总金额（万元）", "持股占比", "备注"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["重交所增资扩股项目成交"][1]],
            ["序号", "项目编号", "标的名称", "投资方名称", "投资金额（万元）", "持股比例", "成交日期", "备注"],
        )
        self.assertEqual(workbook["上海联交所增资项目"].max_row, 1)
        self.assertEqual(workbook["天交所增资扩股项目成交"].max_row, 1)
        self.assertEqual(workbook["重交所增资扩股项目成交"].max_row, 1)

    def test_deal_capital_workbook_preserves_project_investor_order_after_project_sort(self) -> None:
        self._upsert_deal_record(
            record_id="rec-deal-capital-order-b",
            revision_hash="hash-deal-capital-order-b",
            source_id="cbex",
            business_id="deal_capital_increase",
            project_code="G62026BJ000020",
            project_name="后排序增资项目",
            project_type="增资扩股",
            export_extras={
                "investors": [
                    {"name": "投资方B", "amount": "2000", "ratio": "20%"},
                    {"name": "投资方A", "amount": "3000", "ratio": "30%"},
                ],
                "投资总金额（万元）": "5000",
                "持股占比": "50%",
            },
        )
        self._upsert_deal_record(
            record_id="rec-deal-capital-order-a",
            revision_hash="hash-deal-capital-order-a",
            source_id="cbex",
            business_id="deal_capital_increase",
            project_code="G62026BJ000010",
            project_name="先排序增资项目",
            project_type="增资扩股",
            export_extras={
                "investors": [
                    {"name": "投资方C", "amount": "1000", "ratio": "10%"},
                ],
                "投资总金额（万元）": "1000",
                "持股占比": "10%",
            },
        )
        request = ExportRequest(
            date_from="2026-04-18",
            date_to="2026-04-18",
            business_types=["deal_capital_increase"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )

        result = run_ready_export(self.store, request)
        workbook = load_workbook(result.artifacts[0].file_path)
        sheet = workbook["北交所增资项目"]

        self.assertEqual(result.artifacts[0].record_count, 2)
        self.assertEqual([sheet["A2"].value, sheet["A3"].value, sheet["A4"].value], ["G62026BJ000010", "G62026BJ000020", None])
        self.assertEqual([sheet["D2"].value, sheet["D3"].value, sheet["D4"].value], ["投资方C", "投资方B", "投资方A"])

    def test_deal_imputed_date_keeps_remark_suffix(self) -> None:
        self._upsert_deal_record(
            record_id="rec-deal-imputed-date",
            revision_hash="hash-deal-imputed-date",
            source_id="sse",
            business_id="deal_equity_transfer",
            project_code="D32026SH000009",
            project_name="成交日期补全项目",
            project_type="股权转让",
            deal_date="2026-04-20",
            export_extras={
                "备注": "原备注；成交日期缺失，按采集日填列",
            },
        )

        request = ExportRequest(
            date_from="2026-04-20",
            date_to="2026-04-20",
            business_types=["deal_equity_transfer"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["rows"] = rows

        run_ready_export(self.store, request, writer=fake_writer)
        self.assertEqual(len(captured["rows"]), 1)
        row = captured["rows"][0]
        self.assertEqual(row["成交日期"], "2026-04-20")
        self.assertIn("成交日期缺失，按采集日填列", str(row.get("备注") or ""))

    def test_deal_export_uses_collection_date_with_remark_when_real_deal_date_missing(self) -> None:
        self._upsert_deal_record(
            record_id="rec-deal-missing-real-date",
            revision_hash="hash-deal-missing-real-date",
            source_id="sse",
            business_id="deal_equity_transfer",
            project_code="D32026SH000019",
            project_name="缺真实成交日导出项目",
            project_type="股权转让",
            deal_date="",
            collection_date="2026-04-20",
            listing_date="2026-04-20",
            deal_date_basis="collection_date",
            deal_date_is_imputed=True,
            export_extras={"备注": "原备注"},
        )

        request = ExportRequest(
            date_from="2026-04-20",
            date_to="2026-04-20",
            business_types=["deal_equity_transfer"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["rows"] = rows

        run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(len(captured["rows"]), 1)
        row = captured["rows"][0]
        self.assertEqual(row["成交日期"], "2026-04-20")
        self.assertEqual(row["备注"], "原备注；成交日期缺失，按采集日填列")

    def test_deal_capital_workbook_merges_project_level_columns_only(self) -> None:
        for source_id, project_code, project_name in (
            ("cbex", "G62026BJ000002", "北交增资成交项目"),
            ("sse", "G62026SH000002", "上交增资成交项目"),
            ("tpre", "G62026TJ000002", "天交增资成交项目"),
            ("cquae", "G62026CQ000002", "重交增资成交项目"),
        ):
            self._upsert_deal_record(
                record_id=f"rec-deal-capital-merge-{source_id}",
                revision_hash=f"hash-deal-capital-merge-{source_id}",
                source_id=source_id,
                business_id="deal_capital_increase",
                project_code=project_code,
                project_name=project_name,
                project_type="增资扩股",
                deal_price="7200",
                export_extras={
                    "investors": [
                        {"name": "投资方甲", "amount": "4200", "ratio": "58.3%"},
                        {"name": "投资方乙", "amount": "3000", "ratio": "41.7%"},
                    ],
                    "capital_increase_company_name": f"{source_id}融资方",
                    "投资总金额（万元）": "7200",
                    "持股占比（%）": "25%",
                    "持股占比": "25%",
                    "备注": "项目级备注",
                },
            )
        request = ExportRequest(
            date_from="2026-04-18",
            date_to="2026-04-18",
            business_types=["deal_capital_increase"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )

        result = run_ready_export(self.store, request)
        workbook = load_workbook(result.artifacts[0].file_path)
        cbex_ranges = {str(item) for item in workbook["北交所增资项目"].merged_cells.ranges}
        self.assertIn("A2:A3", cbex_ranges)
        self.assertIn("B2:B3", cbex_ranges)
        self.assertIn("C2:C3", cbex_ranges)
        self.assertIn("G2:G3", cbex_ranges)
        self.assertIn("H2:H3", cbex_ranges)
        self.assertIn("I2:I3", cbex_ranges)
        self.assertNotIn("D2:D3", cbex_ranges)
        self.assertNotIn("E2:E3", cbex_ranges)
        self.assertNotIn("F2:F3", cbex_ranges)

        sse_ranges = {str(item) for item in workbook["上海联交所增资项目"].merged_cells.ranges}
        self.assertIn("A2:A3", sse_ranges)
        self.assertIn("B2:B3", sse_ranges)
        self.assertIn("C2:C3", sse_ranges)
        self.assertIn("D2:D3", sse_ranges)
        self.assertIn("H2:H3", sse_ranges)
        self.assertIn("I2:I3", sse_ranges)
        self.assertNotIn("E2:E3", sse_ranges)
        self.assertNotIn("F2:F3", sse_ranges)
        self.assertNotIn("G2:G3", sse_ranges)

        tpre_ranges = {str(item) for item in workbook["天交所增资扩股项目成交"].merged_cells.ranges}
        self.assertIn("A2:A3", tpre_ranges)
        self.assertIn("B2:B3", tpre_ranges)
        self.assertIn("C2:C3", tpre_ranges)
        self.assertIn("G2:G3", tpre_ranges)
        self.assertIn("H2:H3", tpre_ranges)
        self.assertIn("I2:I3", tpre_ranges)
        self.assertNotIn("D2:D3", tpre_ranges)
        self.assertNotIn("E2:E3", tpre_ranges)
        self.assertNotIn("F2:F3", tpre_ranges)

        cquae_ranges = {str(item) for item in workbook["重交所增资扩股项目成交"].merged_cells.ranges}
        self.assertIn("B2:B3", cquae_ranges)
        self.assertIn("C2:C3", cquae_ranges)
        self.assertIn("G2:G3", cquae_ranges)
        self.assertNotIn("A2:A3", cquae_ranges)
        self.assertNotIn("D2:D3", cquae_ranges)
        self.assertNotIn("E2:E3", cquae_ranges)
        self.assertNotIn("F2:F3", cquae_ranges)

    def test_deal_incremental_and_full_respect_new_changed_semantics(self) -> None:
        self._upsert_deal_record(
            record_id="rec-deal-incremental",
            revision_hash="hash-deal-incremental-v1",
            source_id="sse",
            business_id="deal_equity_transfer",
            project_code="D32026SH000099",
            project_name="增量成交项目",
            project_type="股权转让",
        )

        incremental_request = ExportRequest(
            date_from="2026-04-18",
            date_to="2026-04-18",
            business_types=["deal_equity_transfer"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )
        full_request = ExportRequest(
            date_from="2026-04-18",
            date_to="2026-04-18",
            business_types=["deal_equity_transfer"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="deal",
        )

        first = run_ready_export(self.store, incremental_request, writer=lambda path, _rows: self._write_test_export_artifact(path))
        second = run_ready_export(self.store, incremental_request, writer=lambda path, _rows: self._write_test_export_artifact(path))
        self.assertEqual(first.new_records, 1)
        self.assertEqual(first.changed_records, 0)
        self.assertEqual(second.new_records, 0)
        self.assertEqual(second.changed_records, 0)

        self._upsert_deal_record(
            record_id="rec-deal-incremental",
            revision_hash="hash-deal-incremental-v2",
            source_id="sse",
            business_id="deal_equity_transfer",
            project_code="D32026SH000099",
            project_name="增量成交项目",
            project_type="股权转让",
        )
        third = run_ready_export(self.store, incremental_request, writer=lambda path, _rows: self._write_test_export_artifact(path))
        self.assertEqual(third.new_records, 0)
        self.assertEqual(third.changed_records, 1)

        full_first = run_ready_export(self.store, full_request, writer=lambda path, _rows: self._write_test_export_artifact(path))
        full_second = run_ready_export(self.store, full_request, writer=lambda path, _rows: self._write_test_export_artifact(path))
        self.assertEqual(full_first.new_records, 1)
        self.assertEqual(full_first.changed_records, 0)
        self.assertEqual(full_second.new_records, 1)
        self.assertEqual(full_second.changed_records, 0)

    def test_run_ready_export_filters_records_by_record_family(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-deal-export",
                revision_hash="hash-deal-export",
                project_code="D32026SH000002",
                project_name="成交导出项目",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-21",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/deal-export.html",
                archive_path=self._artifact_path("deal-export.html"),
                parser_payload={"项目编号": "D32026SH000002", "项目名称": "成交导出项目"},
                postprocess_payload={"项目编号": "D32026SH000002", "项目名称": "成交导出项目", "项目类型": "股权转让"},
                findings=[],
                record_family="deal",
            )
        )
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="incremental",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="listing",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["file_path"] = file_path
            captured["rows"] = rows

        result = run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(result.new_records, 1)
        self.assertEqual(len(captured["rows"]), 1)
        self.assertEqual(captured["rows"][0]["项目编号"], "G32025SH1000194")

    def test_run_ready_export_filters_records_by_exchange(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-cbex-export",
                revision_hash="hash-cbex-export",
                project_code="G32026BJ1000001",
                project_name="北交所导出项目",
                project_type="股权转让",
                exchange="beijing",
                listing_date="2026-03-21",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/cbex-export.html",
                archive_path=self._artifact_path("cbex-export.html"),
                parser_payload={"项目编号": "G32026BJ1000001", "项目名称": "北交所导出项目"},
                postprocess_payload={"项目编号": "G32026BJ1000001", "项目名称": "北交所导出项目", "项目类型": "股权转让"},
                canonical_record={
                    "record_family": "listing",
                    "canonical_fields": {
                        "project_code": "G32026BJ1000001",
                        "project_name": "北交所导出项目",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "beijing",
                        "start_date": "2026-03-21",
                        "price": "208.00",
                        "seller": "北京测试公司",
                    },
                },
                canonical_projection={
                    "项目编号": "G32026BJ1000001",
                    "项目名称": "北交所导出项目",
                    "项目类型": "股权转让",
                    "项目状态": "挂牌中",
                    "挂牌开始日期": "2026-03-21",
                    "挂牌价格": "208.00",
                    "转让方": "北京测试公司",
                },
                findings=[],
            )
        )
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            exchange="cbex",
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["file_path"] = file_path
            captured["rows"] = rows

        result = run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(result.new_records, 1)
        self.assertEqual(len(captured["rows"]), 1)
        self.assertEqual(captured["rows"][0]["项目编号"], "G32026BJ1000001")

    def test_run_ready_export_filters_exchange_by_source_identity_before_stale_top_level_exchange(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-cbex-source-identity",
                revision_hash="hash-cbex-source-identity",
                project_code="G32026BJ1000002",
                project_name="来源身份北交所项目",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-21",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/cbex-source-identity.html",
                archive_path=self._artifact_path("cbex-source-identity.html"),
                parser_payload={"项目编号": "G32026BJ1000002", "项目名称": "来源身份北交所项目"},
                postprocess_payload={"项目编号": "G32026BJ1000002", "项目名称": "来源身份北交所项目", "项目类型": "股权转让"},
                canonical_record={
                    "record_family": "listing",
                    "source_identity": {"source_id": "cbex"},
                    "business_identity": {
                        "business_id": "equity_transfer",
                        "raw_business_label": "股权转让",
                    },
                    "canonical_fields": {
                        "project_code": "G32026BJ1000002",
                        "project_name": "来源身份北交所项目",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "北交所",
                        "start_date": "2026-03-21",
                        "price": "208.00",
                        "seller": "北京测试公司",
                        "source_type": "国资",
                    },
                },
                findings=[],
                source_identity={
                    "record_family": "listing",
                    "source_id": "cbex",
                    "business_id": "equity_transfer",
                },
            )
        )
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            exchange="cbex",
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["file_path"] = file_path
            captured["rows"] = rows

        result = run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(result.new_records, 1)
        self.assertEqual(len(captured["rows"]), 1)
        self.assertEqual(captured["rows"][0]["项目编号"], "G32026BJ1000002")

    def test_run_ready_export_blocks_corrupt_source_identity_instead_of_exchange_fallback(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-corrupt-source-identity",
                revision_hash="hash-corrupt-source-identity",
                project_code="G32026BJ1000003",
                project_name="来源身份损坏项目",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-21",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/corrupt-source-identity.html",
                archive_path=self._artifact_path("corrupt-source-identity.html"),
                parser_payload={"项目编号": "G32026BJ1000003", "项目名称": "来源身份损坏项目"},
                postprocess_payload={"项目编号": "G32026BJ1000003", "项目名称": "来源身份损坏项目", "项目类型": "股权转让"},
                canonical_record={
                    "record_family": "listing",
                    "source_identity": {"source_id": "cbex"},
                    "business_identity": {
                        "business_id": "equity_transfer",
                        "raw_business_label": "股权转让",
                    },
                    "canonical_fields": {
                        "project_code": "G32026BJ1000003",
                        "project_name": "来源身份损坏项目",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "北交所",
                        "start_date": "2026-03-21",
                        "price": "208.00",
                        "seller": "北京测试公司",
                        "source_type": "国资",
                    },
                },
                findings=[],
                source_identity={
                    "record_family": "listing",
                    "source_id": "cbex",
                    "business_id": "equity_transfer",
                },
            )
        )
        corrupted_canonical_record = {
            "record_family": "listing",
            "source_identity": "oops",
            "business_identity": {
                "business_id": "equity_transfer",
                "raw_business_label": "股权转让",
            },
            "canonical_fields": {
                "project_code": "G32026BJ1000003",
                "project_name": "来源身份损坏项目",
                "project_type": "股权转让",
                "status": "挂牌中",
                "exchange": "北交所",
                "start_date": "2026-03-21",
                "price": "208.00",
                "seller": "北京测试公司",
                "source_type": "国资",
            },
        }
        with self.store._connect() as conn:
            row = conn.execute(
                "SELECT latest_revision_id FROM records WHERE record_id = ?",
                ("rec-corrupt-source-identity",),
            ).fetchone()
            conn.execute(
                "UPDATE records SET source_identity_json = '{}' WHERE record_id = ?",
                ("rec-corrupt-source-identity",),
            )
            conn.execute(
                "UPDATE record_revisions SET canonical_record_json = ? WHERE revision_id = ?",
                (
                    json.dumps(corrupted_canonical_record, ensure_ascii=False, sort_keys=True),
                    int(row["latest_revision_id"]),
                ),
            )
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=None,
            exchange="cbex",
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )

        result = run_ready_export(self.store, request)
        scope_counts = count_records_in_export_scope_by_state(self.store, request)

        self.assertEqual(result.new_records + result.changed_records, 0)
        self.assertEqual(len(result.artifacts), 0)
        self.assertEqual(scope_counts, {"ready": 1})
        self.assertEqual(result.field_missing_blocked_records, 1)
        self.assertEqual(result.field_missing_diagnostics[0]["record_id"], "rec-corrupt-source-identity")
        self.assertEqual(result.field_missing_diagnostics[0]["failure_code"], "invalid_identity_shape")
        self.assertEqual(self.store.list_exports(limit=10), [])
        self.assertEqual(self.store.get_exported_revision_map(result.cursor_id), {})

    def test_run_ready_export_accepts_canonical_business_ids_in_scope(self) -> None:
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["equity_transfer"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
            record_family="listing",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["file_path"] = file_path
            captured["rows"] = rows

        result = run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(result.new_records, 1)
        self.assertEqual(len(captured["rows"]), 1)
        self.assertEqual(captured["rows"][0]["项目编号"], "G32025SH1000194")

    def test_run_ready_export_prefers_business_identity_over_stale_project_type(self) -> None:
        self.store.upsert_record(
            IngestedRecord(
                record_id="rec-physical-business-id",
                revision_hash="hash-physical-business-id",
                project_code="G32025SH1000888",
                project_name="业务身份优先项目",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-21",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/priority.html",
                archive_path=self._artifact_path("priority.html"),
                parser_payload={
                    "项目编号": "G32025SH1000888",
                    "项目名称": "业务身份优先项目",
                    "项目类型": "股权转让",
                },
                postprocess_payload={
                    "项目编号": "G32025SH1000888",
                    "项目名称": "业务身份优先项目",
                    "项目类型": "股权转让",
                },
                canonical_record={
                    "record_family": "listing",
                    "business_identity": {
                        "business_id": "physical_asset",
                        "raw_business_label": "实物资产",
                    },
                    "canonical_fields": {
                        "project_code": "G32025SH1000888",
                        "project_name": "业务身份优先项目",
                        "project_type": "股权转让",
                        "status": "挂牌中",
                        "exchange": "shanghai",
                        "start_date": "2026-03-21",
                        "price": "108.00",
                        "seller": "上海测试公司",
                        "source_type": "国资",
                    },
                },
                canonical_projection={
                    "项目编号": "G32025SH1000888",
                    "项目名称": "业务身份优先项目",
                    "项目类型": "股权转让",
                    "项目状态": "挂牌中",
                    "挂牌开始日期": "2026-03-21",
                    "挂牌价格": "108.00",
                    "转让方": "上海测试公司",
                },
                findings=[],
            )
        )
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["physical_asset"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["file_path"] = file_path
            captured["rows"] = rows

        result = run_ready_export(self.store, request, writer=fake_writer)

        self.assertEqual(result.new_records, 1)
        self.assertEqual(len(captured["rows"]), 1)
        self.assertIn("挂牌_实物资产", str(captured["file_path"]))

    def test_run_ready_export_blocks_malformed_canonical_business_identity(self) -> None:
        class _FakeStore:
            def __init__(self, records: list[dict[str, object]]) -> None:
                self.records = list(records)
                self.marked_records: list[dict[str, object]] = []

            def get_exported_revision_map(self, _cursor_key: str) -> dict[str, dict[str, object]]:
                return {}

            def iter_latest_records(self, **_kwargs) -> list[dict[str, object]]:
                return list(self.records)

            def mark_exported(self, **kwargs) -> None:
                self.marked_records = list(kwargs.get("records") or [])

        fake_store = _FakeStore(
            [
                {
                    "record_id": "rec-top-level-business-id",
                    "revision_id": 1,
                    "revision_hash": "hash-top-level",
                    "project_code": "G32026SH1000101",
                    "project_name": "顶层业务ID优先导出",
                    "project_type": "股权转让",
                    "business_id": "equity_transfer",
                    "record_family": "listing",
                    "exchange": "shanghai",
                    "listing_date": "2026-03-21",
                    "state": "ready",
                    "archive_path": self._artifact_path("top-level-business-id.html"),
                    "canonical_record": {
                        "record_family": "listing",
                        "business_identity": "oops",
                        "canonical_fields": {
                            "project_code": "G32026SH1000101",
                            "project_name": "顶层业务ID优先导出",
                            "project_type": "股权转让",
                            "status": "挂牌中",
                            "exchange": "shanghai",
                            "start_date": "2026-03-21",
                            "price": "120.00",
                            "seller": "上海测试公司",
                        },
                    },
                },
                {
                    "record_id": "rec-missing-business-id",
                    "revision_id": 2,
                    "revision_hash": "hash-missing",
                    "project_code": "G32026SH1000102",
                    "project_name": "缺失业务ID应跳过",
                    "project_type": "股权转让",
                    "business_id": "",
                    "record_family": "listing",
                    "exchange": "shanghai",
                    "listing_date": "2026-03-21",
                    "state": "ready",
                    "canonical_record": {
                        "record_family": "listing",
                        "business_identity": "oops",
                        "canonical_fields": {
                            "project_code": "G32026SH1000102",
                            "project_name": "缺失业务ID应跳过",
                            "project_type": "股权转让",
                            "status": "挂牌中",
                            "exchange": "shanghai",
                            "start_date": "2026-03-21",
                            "price": "99.00",
                            "seller": "上海测试公司",
                        },
                    },
                },
            ]
        )
        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["equity_transfer"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        captured: dict[str, object] = {}

        def fake_writer(file_path: str, rows: list[dict[str, object]]) -> None:
            self._write_test_export_artifact(file_path)
            captured["file_path"] = file_path
            captured["rows"] = rows

        result = run_ready_export(fake_store, request, writer=fake_writer)

        self.assertEqual(result.new_records, 0)
        self.assertEqual(len(result.artifacts), 0)
        self.assertNotIn("rows", captured)
        self.assertEqual(result.field_missing_blocked_records, 1)
        self.assertEqual(
            result.field_missing_diagnostics[0]["record_id"],
            "rec-top-level-business-id",
        )
        self.assertEqual(
            result.field_missing_diagnostics[0]["failure_code"],
            "invalid_identity_shape",
        )
        self.assertEqual(
            [str(item.get("record_id") or "") for item in fake_store.marked_records],
            [],
        )


class StreamingExportRegressionTest(unittest.TestCase):
    """Regression tests for streaming export contract violations."""

    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)

    def _artifact_path(self, name: str, content: bytes = b"fixture artifact") -> str:
        path = os.path.join(self.temp_dir.name, "archive", name)
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "wb") as handle:
            handle.write(content)
        return path

    def _write_test_export_artifact(self, file_path: str) -> None:
        from openpyxl import Workbook

        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        workbook = Workbook()
        workbook.active.append(["ok"])
        workbook.save(file_path)

    def test_export_never_falls_back_to_raw_payload_merge_when_canonical_projection_exists(self) -> None:
        """Regression: streaming_export must never fall back to raw payload merge.

        Once a canonical projection exists, export must use it exclusively.
        Currently export may fall back to merging parser_payload and postprocess_payload
        when canonical_projection is incomplete.
        """
        store = StreamingStore(f"{self.temp_dir.name}/streaming_export_regression.sqlite3", auto_migrate=True)

        # Create a record with canonical_projection that is incomplete
        # but parser_payload has the missing fields
        store.upsert_record(
            IngestedRecord(
                record_id="rec-partial-canonical",
                revision_hash="hash-partial",
                project_code="G32025SH1000999",
                project_name="部分规范化项目",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-21",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/partial.html",
                archive_path=self._artifact_path("partial.html"),
                parser_payload={
                    "项目编号": "G32025SH1000999",
                    "项目名称": "解析层名称",
                    "项目类型": "股权转让",
                    "挂牌价格": "200.00",
                    "转让方": "解析层卖方",
                },
                postprocess_payload={
                    "项目编号": "G32025SH1000999",
                    "项目名称": "后处理名称",
                    "项目类型": "股权转让",
                },
                # canonical_projection is incomplete - missing price and seller
                canonical_projection={
                    "项目编号": "G32025SH1000999",
                    "项目名称": "规范化名称",
                    "项目类型": "股权转让",
                },
                findings=[],
            )
        )

        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        captured_rows = []

        def fake_writer(file_path: str, rows: list[dict]) -> None:
            self._write_test_export_artifact(file_path)
            captured_rows.extend(rows)

        result = run_ready_export(store, request, writer=fake_writer)
        self.assertEqual(result.new_records, 0)

    def test_assemble_normalize_export_preserves_required_canonical_fields(self) -> None:
        """Regression: assemble -> normalize -> export must preserve required canonical fields.

        project_type, status, start_date, price, seller must be preserved.
        """
        store = StreamingStore(f"{self.temp_dir.name}/streaming_export_fields.sqlite3", auto_migrate=True)

        # Create a record with all required canonical fields
        store.upsert_record(
            IngestedRecord(
                record_id="rec-full-canonical",
                revision_hash="hash-full",
                project_code="G32025SH1000194",
                project_name="完整规范化项目",
                project_type="股权转让",
                exchange="shanghai",
                listing_date="2026-03-21",
                state="ready",
                source_file=f"{self.temp_dir.name}/raw/full.html",
                archive_path=self._artifact_path("full.html"),
                parser_payload={
                    "项目编号": "G32025SH1000194",
                    "项目名称": "完整规范化项目",
                    "项目类型": "股权转让",
                    "挂牌开始日期": "2026-03-21",
                    "挂牌价格": "108.00",
                    "转让方": "上海测试公司",
                    "项目状态": "挂牌中",
                },
                postprocess_payload={
                    "项目编号": "G32025SH1000194",
                    "项目名称": "完整规范化项目",
                    "项目类型": "股权转让",
                },
                canonical_record={
                    "record_family": "listing",
                    "canonical_fields": {
                        "project_code": "G32025SH1000194",
                        "project_name": "完整规范化项目",
                        "project_type": "股权转让",
                        "status": "listed",
                        "start_date": "2026-03-21",
                        "price": "108.00",
                        "seller": "上海测试公司",
                    }
                },
                canonical_projection={
                    "项目编号": "G32025SH1000194",
                    "项目名称": "完整规范化项目",
                    "项目类型": "股权转让",
                },
                findings=[],
            )
        )

        request = ExportRequest(
            date_from="2026-03-21",
            date_to="2026-03-21",
            business_types=["股权转让"],
            requested_export_mode="full",
            output_dir=f"{self.temp_dir.name}/exports",
        )
        captured_rows = []

        def fake_writer(file_path: str, rows: list[dict]) -> None:
            self._write_test_export_artifact(file_path)
            captured_rows.extend(rows)

        run_ready_export(store, request, writer=fake_writer)

        exported_row = next(r for r in captured_rows if r.get("项目编号") == "G32025SH1000194")

        # All required canonical fields must be preserved
        # project_type
        self.assertIn("项目类型", exported_row, "project_type must be preserved in export")

        # status (项目状态)
        self.assertIn("项目状态", exported_row, "status must be preserved in export")

        # start_date (挂牌开始日期)
        self.assertIn("挂牌开始日期", exported_row, "start_date must be preserved in export")

        # price (挂牌价格)
        self.assertIn("挂牌价格", exported_row, "price must be preserved in export")

        # seller (转让方)
        self.assertIn("转让方", exported_row, "seller must be preserved in export")


if __name__ == "__main__":
    unittest.main()
