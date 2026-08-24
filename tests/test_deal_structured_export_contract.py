import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

from peap.output_contract import KIND_DEAL_CAPITAL, DealWorkbookSheetSpec
from peap.pipeline_payload_projection import (
    build_export_extras_from_payload,
    normalize_pipeline_payload,
)
from peap.streaming_export import _project_deal_rows, _write_workbook_deal


class _FalsyDict(dict):
    def __bool__(self) -> bool:
        return False


class DealStructuredExportContractTest(unittest.TestCase):
    def test_normalize_pipeline_payload_rejects_explicit_bad_payload_inputs(self) -> None:
        for raw_payload in (False, [], "not-a-payload"):
            with self.subTest(raw_payload=raw_payload):
                with self.assertRaisesRegex(TypeError, "raw_payload"):
                    normalize_pipeline_payload(raw_payload)  # type: ignore[arg-type]

        with self.assertRaisesRegex(TypeError, "standard_payload"):
            normalize_pipeline_payload({"project_code": "P001"}, standard_payload=[])

    def test_normalize_pipeline_payload_preserves_falsy_mapping_payload(self) -> None:
        payload = normalize_pipeline_payload(_FalsyDict({"project_code": "P001"}), standard_payload={})

        self.assertEqual(payload, {"project_code": "P001"})

    def test_normalize_pipeline_payload_honors_explicit_empty_standard_payload(self) -> None:
        payload = normalize_pipeline_payload({"project_code": "P001"}, standard_payload={})

        self.assertEqual(payload, {"project_code": "P001"})

    def test_deal_capital_build_export_extras_synthesizes_investor_from_top_level_fields(self) -> None:
        export_extras = build_export_extras_from_payload(
            {
                "project_code": "G62026SH000120",
                "project_name": "顶层投资方增资项目",
                "deal_date": "2026-04-22",
                "投资方名称": "上海产业投资集团",
                "投资金额（万元）": "12000",
                "持股比例": "12.5%",
            },
            record_family="deal",
            project_type="deal_capital_increase",
            business_id="deal_capital_increase",
        )

        self.assertEqual(
            export_extras.get("investors"),
            [{"name": "上海产业投资集团", "amount": "12000", "ratio": "12.5%"}],
        )

    def test_deal_capital_build_export_extras_synthesizes_investor_from_numeric_top_level_fields(self) -> None:
        export_extras = build_export_extras_from_payload(
            {
                "project_code": "G62026SH000120N",
                "project_name": "顶层投资方数值增资项目",
                "deal_date": "2026-04-22",
                "投资方名称": "上海产业投资集团",
                "投资金额（万元）": 12000.5,
                "持股比例": 12.5,
            },
            record_family="deal",
            project_type="deal_capital_increase",
            business_id="deal_capital_increase",
        )

        self.assertEqual(
            export_extras.get("investors"),
            [{"name": "上海产业投资集团", "amount": "12000.5", "ratio": "12.5"}],
        )

    def test_deal_capital_build_export_extras_replaces_name_only_investors_from_top_level_fields(self) -> None:
        export_extras = build_export_extras_from_payload(
            {
                "project_code": "G62026BJ1000096",
                "project_name": "北交增资成交项目",
                "deal_date": "2026-04-30",
                "investor_name": "北京国泰新华实业有限公司等",
                "investment_amount": "2,800.000000",
                "share_ratio": "11.32",
                "investors": [
                    {"name": "北京国泰新华实业有限公司"},
                    {"name": "郑州国水机械设计研究所有限公司"},
                ],
            },
            record_family="deal",
            project_type="deal_capital_increase",
            business_id="deal_capital_increase",
        )

        self.assertEqual(
            export_extras.get("investors"),
            [{"name": "北京国泰新华实业有限公司等", "amount": "2,800.000000", "ratio": "11.32"}],
        )

    def test_deal_capital_build_export_extras_does_not_synthesize_summary_only_investor(self) -> None:
        export_extras = build_export_extras_from_payload(
            {
                "project_code": "G62026SH000121",
                "project_name": "汇总投资方增资项目",
                "deal_date": "2026-04-22",
                "投资方名称": "总计",
                "投资金额（万元）": "12000",
            },
            record_family="deal",
            project_type="deal_capital_increase",
            business_id="deal_capital_increase",
        )

        self.assertNotIn("investors", export_extras)

    def test_deal_capital_build_export_extras_does_not_synthesize_investor_without_amount(self) -> None:
        export_extras = build_export_extras_from_payload(
            {
                "project_code": "G62026SH000122",
                "project_name": "缺金额增资项目",
                "deal_date": "2026-04-22",
                "投资方名称": "上海产业投资集团",
                "投资金额（万元）": "",
            },
            record_family="deal",
            project_type="deal_capital_increase",
            business_id="deal_capital_increase",
        )

        self.assertNotIn("investors", export_extras)

    def test_deal_capital_build_export_extras_does_not_synthesize_investor_with_malformed_amount(self) -> None:
        export_extras = build_export_extras_from_payload(
            {
                "project_code": "G62026SH000122A",
                "project_name": "异常金额增资项目",
                "deal_date": "2026-04-22",
                "投资方名称": "上海产业投资集团",
                "投资金额（万元）": {"bad": "data"},
            },
            record_family="deal",
            project_type="deal_capital_increase",
            business_id="deal_capital_increase",
        )

        self.assertNotIn("investors", export_extras)

    def test_deal_investor_expansion_uses_projection_output_kind_not_business_id(self) -> None:
        record = {
            "record_family": "deal",
            "source_identity": {"source_id": "sse"},
            "canonical_record": {
                "record_family": "deal",
                "source_identity": {"source_id": "sse"},
                "business_identity": {"business_id": "deal_capital_increase"},
                "canonical_fields": {
                    "project_code": "G62026SH000001",
                    "project_name": "增资成交项目",
                    "status": "已成交",
                    "deal_date": "2026-04-20",
                },
                "export_extras": {
                    "investors": [
                        {"name": "投资方甲", "amount": "300"},
                        {"name": "投资方乙", "amount": "500"},
                    ],
                },
            },
        }
        payload = {"项目编号": "G62026SH000001", "项目名称": "增资成交项目", "成交日期": "2026-04-20"}

        def fake_extra_fields(kind: str) -> list[str]:
            return ["investors"] if kind == "deal_capital_output_v2" else []

        with patch("peap.streaming_export.get_structured_export_extra_fields", side_effect=fake_extra_fields):
            rows = _project_deal_rows(
                record,
                payload,
                business_id="deal_capital_increase",
                output_kind="deal_capital_output_v2",
            )

        self.assertEqual([row["投资方名称"] for row in rows], ["投资方甲", "投资方乙"])

    def test_deal_workbook_contract_uses_output_kind_not_business_id(self) -> None:
        rows = [
            {
                "_source_id": "sse",
                "_group_project_code": "G62026SH000001",
                "_group_deal_date": "2026-04-20",
                "项目编号": "G62026SH000001",
                "项目名称": "增资成交项目",
                "成交日期": "2026-04-20",
                "投资方名称": "投资方甲",
            }
        ]
        seen_kinds: list[str] = []

        def fake_sheet_specs(kind: str) -> list[DealWorkbookSheetSpec]:
            seen_kinds.append(kind)
            if kind == "deal_capital_output_v2":
                return [
                    DealWorkbookSheetSpec(
                        source_id="sse",
                        sheet_name="custom-capital",
                        headers=("项目编号", "投资方名称"),
                    )
                ]
            return []

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as workbook_file:
            with patch("peap.streaming_export.list_deal_workbook_sheet_specs", side_effect=fake_sheet_specs):
                _write_workbook_deal(
                    workbook_file.name,
                    rows,
                    output_kind="deal_capital_output_v2",
                )

            workbook = load_workbook(workbook_file.name)

        self.assertEqual(seen_kinds, ["deal_capital_output_v2"])
        self.assertEqual(workbook.sheetnames, ["custom-capital"])
        self.assertEqual([cell.value for cell in workbook["custom-capital"][1]], ["项目编号", "投资方名称"])

    def test_deal_capital_standard_only_project_fields_reach_workbook_columns(self) -> None:
        export_extras = build_export_extras_from_payload(
            {
                "project_code": "G62026SH000099",
                "project_name": "标准字段增资成交项目",
                "deal_date": "2026-04-22",
                "capital_company_name": "上海标准增资企业",
                "total_investment_amount": "5000",
                "holding_ratio": "35%",
            },
            record_family="deal",
            project_type="deal_capital_increase",
            business_id="deal_capital_increase",
        )

        self.assertEqual(export_extras.get("增资企业名称"), "上海标准增资企业")
        self.assertEqual(export_extras.get("投资总金额（万元）"), "5000")
        self.assertEqual(export_extras.get("持股占比"), "35%")

        record = {
            "record_family": "deal",
            "source_identity": {"source_id": "sse"},
            "canonical_record": {
                "record_family": "deal",
                "source_identity": {"source_id": "sse"},
                "business_identity": {"business_id": "deal_capital_increase"},
                "canonical_fields": {
                    "project_code": "G62026SH000099",
                    "project_name": "标准字段增资成交项目",
                    "status": "已成交",
                    "deal_date": "2026-04-22",
                },
                "export_extras": export_extras,
            },
        }
        payload = {"项目编号": "G62026SH000099", "项目名称": "标准字段增资成交项目", "成交日期": "2026-04-22"}
        rows = _project_deal_rows(
            record,
            payload,
            business_id="deal_capital_increase",
            output_kind=KIND_DEAL_CAPITAL,
        )

        self.assertEqual(rows[0]["增资企业名称"], "上海标准增资企业")
        self.assertEqual(rows[0]["投资总金额（万元）"], "5000")
        self.assertEqual(rows[0]["持股占比"], "35%")

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as workbook_file:
            _write_workbook_deal(workbook_file.name, rows, output_kind=KIND_DEAL_CAPITAL)
            workbook = load_workbook(workbook_file.name)

        sheet = workbook["上海联交所增资项目"]
        headers = [cell.value for cell in sheet[1]]
        values = [cell.value for cell in sheet[2]]
        row = dict(zip(headers, values, strict=False))
        self.assertEqual(row["增资企业名称"], "上海标准增资企业")
        self.assertEqual(row["投资总金额（万元）"], "5000")
        self.assertEqual(row["持股占比"], "35%")


if __name__ == "__main__":
    unittest.main()
