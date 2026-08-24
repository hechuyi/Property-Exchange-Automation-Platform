"""Export ready records from the streaming store."""

from __future__ import annotations

import datetime as dt
import hashlib
import json
import os
import re
import stat
import uuid
from collections import defaultdict
from collections.abc import Mapping
from typing import Any, Dict, Iterable, List

from peap_core.business_catalog import resolve_business_descriptor
from peap_core.family_catalog import get_family_descriptor
from peap_core.field_missing_contract import normalize_missing_fields
from peap_core.source_catalog import canonical_source_code

from .artifact_truth import resolve_artifact_evidence_verdict
from .export_evidence_policy import export_evidence_verdict_accepted
from .export_projection import (
    ExportProjectionError,
    project_canonical_record_to_export_payload,
)
from .export_projection import (
    is_export_empty_value as _is_export_empty_value,
)
from .output_contract import (
    DEAL_SOURCE_ORDER,
    KIND_DEAL_CAPITAL,
    clone_field_candidates,
    get_output_columns_for_kind,
    get_structured_export_extra_fields,
    list_deal_workbook_sheet_specs,
)
from .projection_registry import list_projection_profiles, resolve_projection_profile
from .streaming_models import ExportArtifact, ExportRequest, ExportRunResult, PostProcessFinding
from .streaming_postprocess import is_summary_investor_name
from .streaming_store import StreamingStore

BUSINESS_TYPE_LABELS = {
    "股权转让": "挂牌_股权转让",
    "实物资产": "挂牌_实物资产",
    "增资扩股": "挂牌_增资扩股",
    "预披露": "挂牌_预披露",
}

HEADER_PRIORITY = [
    "项目编号",
    "项目名称",
    "项目类型",
    "项目状态",
    "交易所",
    "类型",
    "转让方",
    "融资方",
    "隶属集团",
    "挂牌开始日期",
    "挂牌截止日期",
    "披露开始日期",
    "披露截止日期",
    "预披露开始日期",
    "预披露截止日期",
    "挂牌价格",
    "融资金额",
    "受让方名称",
    "备注",
]

DEAL_CAPITAL_INVESTOR_HEADERS = frozenset({"投资方名称", "投资金额（万元）", "持股比例", "持股比例（%）"})
IMPUTED_DEAL_DATE_REMARK_SUFFIX = "成交日期缺失，按采集日填列"
def _export_cell_value(value: Any) -> Any:
    return "" if _is_export_empty_value(value) else value


def _safe_suffix(value: str) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|]+", "_", str(value or "").strip())
    return cleaned or "default"


def _normalize_export_mode(raw_value: str) -> str:
    text = str(raw_value or "full").strip().lower()
    if text == "full":
        return "full"
    if text == "incremental":
        return "incremental"
    raise ValueError("requested_export_mode must be full or incremental")


def _normalize_requested_state(raw_value: str) -> str:
    return str(raw_value or "all").strip().lower() or "all"


def _normalize_keyword(raw_value: str) -> str:
    return str(raw_value or "").strip().lower()


def _normalize_exchange(raw_value: str) -> str:
    return str(canonical_source_code(raw_value or "all") or "all").strip() or "all"


def _request_record_family(request: ExportRequest) -> str:
    raw_family = str(request.record_family or "").strip()
    if not raw_family:
        return "listing"
    return get_family_descriptor(raw_family).family_id


def _normalized_business_scope_ids(
    business_types: list[str] | None,
    *,
    record_family: str,
) -> list[str]:
    if business_types is None:
        raw_business_types: list[str] = []
    elif isinstance(business_types, Mapping) or isinstance(business_types, (str, bytes)):
        raise TypeError("business_types must be a list of strings")
    else:
        try:
            raw_business_types = list(business_types)
        except TypeError:
            raise TypeError("business_types must be a list of strings") from None
    normalized: list[str] = []
    seen: set[str] = set()
    for raw_value in raw_business_types:
        text = str(raw_value or "").strip()
        if not text:
            continue
        profile = resolve_projection_profile(record_family, text)
        if profile is not None:
            business_id = profile.business_id
        else:
            try:
                descriptor = resolve_business_descriptor(text, family_id=record_family)
            except KeyError:
                descriptor = None
            business_id = descriptor.business_id if descriptor is not None else text
        if business_id in seen:
            continue
        seen.add(business_id)
        normalized.append(business_id)
    if normalized:
        return normalized
    return [profile.business_id for profile in list_projection_profiles(record_family=record_family)]


def _record_findings_list(record: Dict[str, Any]) -> list[Any]:
    findings = record.get("findings")
    if findings is None:
        return []
    if isinstance(findings, Mapping) or isinstance(findings, (str, bytes)):
        raise TypeError("record.findings must be a list of findings")
    try:
        items = list(findings)
    except TypeError:
        raise TypeError("record.findings must be a list of findings") from None
    for item in items:
        if not isinstance(item, Mapping):
            raise TypeError("record.findings[*] must be a mapping")
    return items


def _missing_fields_argument_list(missing_fields: Iterable[Mapping[str, Any]] | None) -> list[dict[str, str]]:
    if missing_fields is None:
        return []
    if isinstance(missing_fields, Mapping) or isinstance(missing_fields, (str, bytes)):
        raise TypeError("missing_fields must be a list of mappings")
    try:
        items = list(missing_fields)
    except TypeError:
        raise TypeError("missing_fields must be a list of mappings") from None
    for item in items:
        if not isinstance(item, Mapping):
            raise TypeError("missing_fields[*] must be a mapping")
    return [dict(item) for item in items]


def _projection_error_missing_fields(exc: ExportProjectionError) -> list[dict[str, str]]:
    return _missing_fields_argument_list(getattr(exc, "missing_fields", None))


def _projection_failure_finding(exc: ExportProjectionError) -> PostProcessFinding | None:
    failure_code = str(getattr(exc, "failure_code", "") or "").strip()
    if failure_code not in {"canonical_field_missing", "export_field_missing"}:
        return None
    structured_missing_fields = _projection_error_missing_fields(exc)
    if failure_code == "canonical_field_missing":
        missing_fields: list[Any] = [
            str(item.get("canonical_field") or item.get("field") or "").strip()
            for item in structured_missing_fields
            if str(item.get("canonical_field") or item.get("field") or "").strip()
        ]
    else:
        missing_fields = [
            {
                "export_field": str(item.get("export_field") or item.get("field") or "").strip(),
                "canonical_field": str(item.get("canonical_field") or "").strip(),
            }
            for item in structured_missing_fields
            if str(item.get("export_field") or item.get("field") or item.get("canonical_field") or "").strip()
        ]
    return PostProcessFinding(
        severity="warn",
        type=failure_code,
        message=str(exc),
        evidence={"missing_fields": missing_fields},
    )


def _demote_ready_projection_failure(
    store: StreamingStore,
    record: Mapping[str, Any],
    finding: PostProcessFinding,
) -> None:
    """Persist a projection failure when the store supports state repair."""
    demoter = getattr(store, "mark_ready_record_field_missing", None)
    if not callable(demoter):
        return
    demoter(
        str(record.get("record_id") or ""),
        revision_id=int(record.get("revision_id") or 0),
        finding=finding,
    )


def _record_business_id(record: Dict[str, Any]) -> str:
    direct = str(record.get("business_id") or "").strip()
    if direct:
        return direct
    canonical_record = record.get("canonical_record")
    raw_business_identity = canonical_record.get("business_identity") if isinstance(canonical_record, dict) else None
    if raw_business_identity is not None and not isinstance(raw_business_identity, Mapping):
        raise ExportProjectionError(
            f"canonical_record.business_identity must be an object when present, got {type(raw_business_identity).__name__}",
            failure_code="invalid_identity_shape",
        )
    business_identity = dict(raw_business_identity) if isinstance(raw_business_identity, Mapping) else {}
    nested = str(business_identity.get("business_id") or "").strip()
    if nested:
        return nested
    raw_source_identity = canonical_record.get("source_identity") if isinstance(canonical_record, dict) else None
    if raw_source_identity is not None and not isinstance(raw_source_identity, Mapping):
        raise ExportProjectionError(
            f"canonical_record.source_identity must be an object when present, got {type(raw_source_identity).__name__}",
            failure_code="invalid_identity_shape",
        )
    source_identity = dict(raw_source_identity) if isinstance(raw_source_identity, Mapping) else {}
    source_business_id = str(source_identity.get("business_id") or "").strip()
    if source_business_id:
        return source_business_id
    return ""


def _record_business_id_for_diagnostic(record: Dict[str, Any]) -> str:
    try:
        return _record_business_id(record)
    except ExportProjectionError:
        return str(record.get("business_id") or "").strip()


def _record_source_id(record: Dict[str, Any]) -> str:
    source_identity = record.get("source_identity_json")
    if not isinstance(source_identity, dict):
        legacy_source_identity = record.get("source_identity")
        if legacy_source_identity is None:
            source_identity = {}
        elif isinstance(legacy_source_identity, Mapping):
            source_identity = dict(legacy_source_identity)
        else:
            raise TypeError(
                f"source_identity must be a mapping when present, got {type(legacy_source_identity).__name__}"
            )
    source_id = str(source_identity.get("source_id") or "").strip()
    if source_id:
        return _normalize_exchange(source_id)
    canonical_record = record.get("canonical_record")
    raw_canonical_source_identity = canonical_record.get("source_identity") if isinstance(canonical_record, dict) else None
    if raw_canonical_source_identity is not None and not isinstance(raw_canonical_source_identity, Mapping):
        raise ExportProjectionError(
            f"canonical_record.source_identity must be an object when present, got {type(raw_canonical_source_identity).__name__}",
            failure_code="invalid_identity_shape",
        )
    canonical_source_identity = (
        dict(raw_canonical_source_identity)
        if isinstance(raw_canonical_source_identity, Mapping)
        else {}
    )
    source_id = str(canonical_source_identity.get("source_id") or "").strip()
    if source_id:
        return _normalize_exchange(source_id)
    if _record_family_of_record(record) == "deal":
        raise ExportProjectionError(
            "deal record source_identity.source_id is required for export source routing",
            failure_code="source_identity_missing",
            missing_fields=[
                {
                    "kind": "canonical",
                    "field": "source_identity.source_id",
                    "canonical_field": "source_identity.source_id",
                    "export_field": "交易所",
                    "message": "source_identity.source_id is required for deal export source routing",
                }
            ],
        )
    return _normalize_exchange(str(record.get("exchange") or ""))


def _record_family_of_record(record: Dict[str, Any]) -> str:
    direct = str(record.get("record_family") or "").strip()
    if direct:
        return direct
    canonical_record = record.get("canonical_record")
    if isinstance(canonical_record, dict):
        nested = str(canonical_record.get("record_family") or "").strip()
        if nested:
            return nested
        business_identity = canonical_record.get("business_identity")
        if isinstance(business_identity, dict):
            nested = str(business_identity.get("record_family") or "").strip()
            if nested:
                return nested
        source_identity = canonical_record.get("source_identity")
        if isinstance(source_identity, dict):
            nested = str(source_identity.get("record_family") or "").strip()
            if nested:
                return nested
    return "listing"


def _first_non_empty_value(payload: Dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = payload.get(key)
        if not _is_export_empty_value(value):
            return str(value).strip()
    return ""


def _join_name_list(values: object) -> str:
    if isinstance(values, list):
        out = [str(item).strip() for item in values if str(item).strip()]
        return "；".join(out)
    if values is None:
        return ""
    return str(values).strip()


def _is_summary_investor(name: str) -> bool:
    normalized = str(name or "").strip()
    if not normalized:
        return False
    return bool(is_summary_investor_name(normalized))


def _ensure_supported_deal_source(source_id: str) -> str:
    normalized = _normalize_exchange(source_id)
    if normalized in DEAL_SOURCE_ORDER:
        return normalized
    allowed = ", ".join(DEAL_SOURCE_ORDER)
    raise ExportProjectionError(
        f"unsupported deal source_id for export: {normalized!r}; expected one of: {allowed}"
    )


def _normalize_investor_entries(values: object) -> list[Dict[str, str]]:
    if not isinstance(values, list):
        return []
    entries: list[Dict[str, str]] = []
    for raw_entry in values:
        if isinstance(raw_entry, dict):
            entry = dict(raw_entry)
        else:
            entry = {"name": raw_entry}
        name = _first_non_empty_value(entry, "name", "投资方名称", "投资方", "投资人", "investor_name")
        if _is_summary_investor(name):
            continue
        amount = _first_non_empty_value(
            entry,
            "amount",
            "投资金额（万元）",
            "投资金额",
            "投资额",
            "investment_amount",
            "investmentAmount",
        )
        ratio = _first_non_empty_value(
            entry,
            "ratio",
            "持股比例",
            "持股比例（%）",
            "投资比例",
            "持股占比",
            "持股占比（%）",
        )
        entries.append(
            {
                "投资方名称": name,
                "投资金额（万元）": amount,
                "持股比例": ratio,
                "持股比例（%）": ratio,
            }
        )
    return entries


def _sort_int(value: Any) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _deal_canonical_fields(record: Dict[str, Any]) -> Dict[str, Any]:
    canonical_record = record.get("canonical_record")
    if not isinstance(canonical_record, dict):
        return {}
    canonical_fields = canonical_record.get("canonical_fields")
    return dict(canonical_fields) if isinstance(canonical_fields, dict) else {}


def _coerce_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"true", "1", "yes", "y"}


def _append_remark_suffix(row: Dict[str, Any], suffix: str) -> None:
    if not suffix:
        return
    remark = str(row.get("备注") or "").strip()
    if suffix in remark:
        row["备注"] = remark
        return
    row["备注"] = f"{remark}；{suffix}" if remark else suffix


def _apply_deal_date_export_semantics(row: Dict[str, Any], record: Dict[str, Any]) -> None:
    canonical_fields = _deal_canonical_fields(record)
    basis = str(canonical_fields.get("deal_date_basis") or "").strip()
    collection_date = str(canonical_fields.get("collection_date") or "").strip()
    deal_date_is_imputed = _coerce_bool(canonical_fields.get("deal_date_is_imputed"))
    if basis != "collection_date" and not deal_date_is_imputed:
        return
    if not collection_date:
        return
    if not str(row.get("成交日期") or "").strip():
        row["成交日期"] = collection_date
    if str(row.get("成交日期") or "").strip() == collection_date:
        _append_remark_suffix(row, IMPUTED_DEAL_DATE_REMARK_SUFFIX)


def _deal_row_from_payload(
    payload: Dict[str, Any],
    *,
    record: Dict[str, Any],
    source_id: str,
    business_id: str,
) -> Dict[str, Any]:
    row = dict(payload)
    _apply_deal_date_export_semantics(row, record)
    row["标的名称"] = _first_non_empty_value(row, "标的名称", "项目名称")
    row["转让标的评估值"] = _first_non_empty_value(row, "转让标的评估值", "转让标的评估结果")
    row["转让标的评估结果"] = _first_non_empty_value(row, "转让标的评估结果", "转让标的评估值")
    reserve_price = _first_non_empty_value(row, "转让底价", "转让底价（万元）")
    row["转让底价"] = reserve_price
    row["转让底价（万元）"] = _first_non_empty_value(row, "转让底价（万元）", "转让底价")
    deal_price = _first_non_empty_value(row, "交易价格", "交易价格（万元）", "成交金额")
    row["交易价格"] = deal_price
    row["交易价格（万元）"] = _first_non_empty_value(row, "交易价格（万元）", "交易价格", "成交金额")
    row["_source_id"] = source_id
    row["_business_id"] = business_id
    row["_group_project_code"] = _first_non_empty_value(row, "项目编号")
    row["_group_deal_date"] = _first_non_empty_value(row, "成交日期")
    return row


def _project_deal_rows(
    record: Dict[str, Any],
    payload: Dict[str, Any],
    *,
    business_id: str,
    output_kind: str,
) -> list[Dict[str, Any]]:
    source_id = _ensure_supported_deal_source(_record_source_id(record))
    row = _deal_row_from_payload(payload, record=record, source_id=source_id, business_id=business_id)
    canonical_record = record.get("canonical_record")
    raw_export_extras = canonical_record.get("export_extras") if isinstance(canonical_record, dict) else None
    if raw_export_extras is None:
        export_extras = {}
    elif isinstance(raw_export_extras, Mapping):
        export_extras = dict(raw_export_extras)
    else:
        raise TypeError(
            f"canonical_record.export_extras must be a mapping when present, "
            f"got {type(raw_export_extras).__name__}"
        )

    if "investors" not in set(get_structured_export_extra_fields(output_kind)):
        return [row]

    financing_names = export_extras.get("financing_party_names")
    financing_name = _join_name_list(financing_names)
    if not financing_name:
        financing_name = _first_non_empty_value(
            export_extras,
            "capital_increase_company_name",
            "增资企业名称",
            "融资方",
            "融资方名称",
        )
    row["增资企业名称"] = financing_name
    row["投资总金额（万元）"] = _first_non_empty_value(
        export_extras,
        "投资总金额（万元）",
        "投资总金额",
        "融资金额",
        "融资金额（万元）",
    )
    holding_ratio = _first_non_empty_value(export_extras, "持股占比（%）", "持股占比", "持股比例")
    row["持股占比（%）"] = holding_ratio
    row["持股占比"] = _first_non_empty_value(export_extras, "持股占比", "持股占比（%）", "持股比例")

    raw_investors = export_extras.get("investors")
    if "investors" in export_extras and not isinstance(raw_investors, list):
        raise ExportProjectionError(
            f"canonical_record.export_extras.investors must be a list when present, "
            f"got {type(raw_investors).__name__}"
        )
    investors = _normalize_investor_entries(raw_investors)
    if not investors:
        investors = [{"投资方名称": "", "投资金额（万元）": "", "持股比例": "", "持股比例（%）": ""}]

    expanded_rows: list[Dict[str, Any]] = []
    for investor_order, investor_entry in enumerate(investors):
        expanded = dict(row)
        expanded.update(investor_entry)
        expanded["_investor_order"] = investor_order
        for key in ("investors", "transferors", "financing_party_names", "project_parties"):
            value = export_extras.get(key)
            if isinstance(value, list):
                expanded[key] = list(value)
        expanded_rows.append(expanded)
    return expanded_rows


def _project_export_rows(
    record: Dict[str, Any],
    payload: Dict[str, Any],
    *,
    record_family: str,
    business_id: str,
    output_kind: str,
) -> list[Dict[str, Any]]:
    if record_family == "deal":
        return _project_deal_rows(record, payload, business_id=business_id, output_kind=output_kind)
    return [payload]


def _record_matches_keyword(record: Dict[str, Any], *, keyword: str) -> bool:
    normalized_keyword = _normalize_keyword(keyword)
    if not normalized_keyword:
        return True
    try:
        payload = record_to_export_payload(record)
    except ExportProjectionError:
        payload = {}
        canonical_record = record.get("canonical_record")
        canonical_fields = canonical_record.get("canonical_fields") if isinstance(canonical_record, Mapping) else None
        if isinstance(canonical_fields, Mapping):
            payload.update({str(key): value for key, value in canonical_fields.items()})
        canonical_projection = record.get("canonical_projection")
        if isinstance(canonical_projection, Mapping):
            payload.update({str(key): value for key, value in canonical_projection.items()})
    search_blob = " ".join(
        [
            str(record.get("project_code") or ""),
            str(record.get("project_name") or ""),
            str(record.get("project_type") or ""),
            str(record.get("exchange") or ""),
            str(record.get("listing_date") or ""),
            str(record.get("state") or ""),
        ]
        + [str(value or "") for value in payload.values()]
    ).lower()
    return normalized_keyword in search_blob


def _canonical_scope_hash(request: ExportRequest, *, record_family: str | None = None) -> str:
    scope = _export_scope_dto(request, record_family=record_family)
    seed = json.dumps(scope, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha1(seed.encode("utf-8")).hexdigest()[:12]


def _export_scope_dto(request: ExportRequest, *, record_family: str | None = None) -> dict[str, Any]:
    requested_state = _normalize_requested_state(request.requested_state)
    keyword = _normalize_keyword(request.keyword)
    exchange = _normalize_exchange(request.exchange)
    record_family = record_family or _request_record_family(request)
    business_types = _normalized_business_scope_ids(
        request.business_types,
        record_family=record_family,
    )
    return {
        "record_family": record_family,
        "date_from": request.date_from or "",
        "date_to": request.date_to or "",
        "business_ids": sorted(business_types),
        "exchange": exchange,
        "requested_state": requested_state,
        "keyword": keyword,
    }


def _export_profile_id_for_scope(
    request: ExportRequest,
    business_types: list[str] | None = None,
    *,
    record_family: str | None = None,
) -> str:
    record_family = record_family or _request_record_family(request)
    resolved_business_types = list(
        business_types
        if business_types is not None
        else _normalized_business_scope_ids(request.business_types, record_family=record_family)
    )
    if len(resolved_business_types) == 1:
        profile = resolve_projection_profile(record_family, resolved_business_types[0])
        if profile is not None:
            return profile.profile_id
    seed = ",".join(sorted(resolved_business_types)) or "all"
    return f"{record_family}/{seed}"


def _default_cursor_id(request: ExportRequest, *, record_family: str | None = None) -> str:
    profile_id = _export_profile_id_for_scope(request, record_family=record_family).replace("/", "-")
    return f"ready_export:profile-{profile_id}:{_canonical_scope_hash(request, record_family=record_family)}:schema-v1:headers-v1"


def _business_scope_covers_family(record_family: str, business_types: list[str]) -> bool:
    family_business_ids = {profile.business_id for profile in list_projection_profiles(record_family=record_family)}
    return bool(family_business_ids) and set(business_types) == family_business_ids


def _iter_records_in_export_scope(
    store: StreamingStore,
    request: ExportRequest,
    *,
    include_all_states: bool,
    record_family: str | None = None,
) -> List[Dict[str, Any]]:
    record_family = record_family or _request_record_family(request)
    requested_state = _normalize_requested_state(request.requested_state)
    keyword = _normalize_keyword(request.keyword)
    requested_exchange = _normalize_exchange(request.exchange)
    business_types = _normalized_business_scope_ids(
        request.business_types,
        record_family=record_family,
    )
    family_wide_business_scope = _business_scope_covers_family(record_family, business_types)
    records = store.iter_latest_records(
        states=None if include_all_states else ["ready"],
        date_from=request.date_from,
        date_to=request.date_to,
        record_family=record_family,
    )

    scoped: List[Dict[str, Any]] = []
    for record in records:
        record_state = str(record.get("state") or "").strip().lower()
        if record_state == "field_missing" and not include_all_states:
            continue
        if requested_state not in {"", "all"} and record_state != requested_state:
            continue
        try:
            record_exchange = _record_source_id(record)
        except ExportProjectionError:
            record_exchange = ""
        if requested_exchange not in {"", "all"} and record_exchange and record_exchange != requested_exchange:
            continue
        try:
            business_id = _record_business_id(record)
        except ExportProjectionError:
            business_id = ""
            if not family_wide_business_scope:
                continue
        if business_types and business_id not in business_types and not family_wide_business_scope:
            continue
        if not _record_matches_keyword(record, keyword=keyword):
            continue
        scoped.append(record)
    return scoped


def count_records_in_export_scope_by_state(store: StreamingStore, request: ExportRequest) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for record in _iter_records_in_export_scope(store, request, include_all_states=True):
        state = str(record.get("state") or "").strip()
        if not state:
            continue
        counts[state] = counts.get(state, 0) + 1
    return counts


def _field_missing_diagnostic(
    record: Dict[str, Any],
    *,
    failure_code: str = "",
    missing_fields: list[dict[str, str]] | None = None,
    evidence_status: str = "",
    evidence_reason_code: str = "",
    message: str = "",
) -> Dict[str, Any]:
    raw_missing_fields: list[Any] = []
    raw_missing_fields.extend(_missing_fields_argument_list(missing_fields))
    projection_finding_types: list[str] = []
    for finding in _record_findings_list(record):
        finding_type = str(finding.get("type") or "").strip()
        if finding_type in {"canonical_field_missing", "export_field_missing"}:
            projection_finding_types.append(finding_type)
        evidence = finding.get("evidence") if isinstance(finding, dict) else None
        if not isinstance(evidence, dict):
            continue
        if "missing_fields" not in evidence:
            continue
        evidence_missing_fields = evidence.get("missing_fields")
        if isinstance(evidence_missing_fields, dict):
            raw_missing_fields.append(evidence_missing_fields)
            continue
        if not isinstance(evidence_missing_fields, list | tuple):
            raise TypeError(
                "field_missing finding evidence.missing_fields must be a list, tuple, or dict"
            )
        for field_name in evidence_missing_fields:
            raw_missing_fields.append(field_name)
    resolved_failure_code = str(failure_code or "").strip()
    if not resolved_failure_code:
        resolved_failure_code = (
            "canonical_field_missing"
            if "canonical_field_missing" in projection_finding_types
            else "export_field_missing"
        )
    diagnostic = {
        "record_id": str(record.get("record_id") or ""),
        "revision_id": int(record.get("revision_id") or 0),
        "record_family": _record_family_of_record(record),
        "business_id": _record_business_id_for_diagnostic(record),
        "project_code": str(record.get("project_code") or ""),
        "project_name": str(record.get("project_name") or ""),
        "failure_code": resolved_failure_code,
        "missing_fields": normalize_missing_fields(raw_missing_fields),
    }
    if evidence_status:
        diagnostic["evidence_status"] = evidence_status
    if evidence_reason_code:
        diagnostic["evidence_reason_code"] = evidence_reason_code
    if message:
        diagnostic["message"] = message
    return diagnostic


def _artifact_evidence_rejection_message(status: str, reason_code: str) -> str:
    if status == "identity_mismatch" and reason_code == "project_code_mismatch":
        return "本地证据身份与数据库项目编号不一致，已阻止导出"
    if status:
        return f"本地证据未通过导出校验：{status}"
    return "本地证据未通过导出校验"


def _artifact_checksums(artifacts: list[ExportArtifact]) -> dict[str, str]:
    checksums: dict[str, str] = {}
    for index, artifact in enumerate(artifacts):
        path = str(artifact.file_path or "").strip()
        if not path:
            raise ExportProjectionError(
                f"export artifact path is empty at index {index}",
                failure_code="export_artifact_missing",
            )
        if path in checksums:
            raise ExportProjectionError(
                f"duplicate export artifact path: {path}",
                failure_code="export_artifact_invalid",
            )
        try:
            path_stat = os.lstat(path)
            if stat.S_ISLNK(path_stat.st_mode) or not stat.S_ISREG(path_stat.st_mode):
                raise ExportProjectionError(
                    f"export artifact is not a regular file: {path}",
                    failure_code="export_artifact_invalid",
                )
            flags = os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0)
            fd = os.open(path, flags)
            try:
                opened_stat = os.fstat(fd)
                if not stat.S_ISREG(opened_stat.st_mode) or not os.path.samestat(path_stat, opened_stat):
                    raise ExportProjectionError(
                        f"export artifact changed before checksum: {path}",
                        failure_code="export_artifact_invalid",
                    )
                digest = hashlib.sha256()
                with os.fdopen(fd, "rb", closefd=False) as handle:
                    for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                        digest.update(chunk)
                final_stat = os.fstat(fd)
            finally:
                os.close(fd)
            current_stat = os.lstat(path)
            if (
                not os.path.samestat(opened_stat, final_stat)
                or not os.path.samestat(final_stat, current_stat)
                or final_stat.st_size != opened_stat.st_size
                or final_stat.st_mtime_ns != opened_stat.st_mtime_ns
                or final_stat.st_nlink < 1
            ):
                raise ExportProjectionError(
                    f"export artifact changed during checksum: {path}",
                    failure_code="export_artifact_invalid",
                )
        except ExportProjectionError:
            raise
        except FileNotFoundError as exc:
            raise ExportProjectionError(
                f"export artifact is missing: {path}",
                failure_code="export_artifact_missing",
            ) from exc
        except OSError as exc:
            raise ExportProjectionError(
                f"export artifact cannot be read: {path}: {exc}",
                failure_code="export_artifact_invalid",
            ) from exc
        checksums[path] = digest.hexdigest()
    return checksums


def _ensure_valid_xlsx_artifact(file_path: str) -> None:
    if not os.path.isfile(file_path):
        raise ExportProjectionError(
            f"export writer did not create artifact: {file_path}",
            failure_code="export_artifact_missing",
        )
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, read_only=True, data_only=False)
        try:
            if not workbook.sheetnames:
                raise ExportProjectionError(
                    f"export writer created invalid artifact without sheets: {file_path}",
                    failure_code="export_artifact_invalid",
                )
        finally:
            workbook.close()
    except ExportProjectionError:
        raise
    except Exception as exc:
        raise ExportProjectionError(
            f"export writer created invalid xlsx artifact: {file_path}",
            failure_code="export_artifact_invalid",
        ) from exc


def _remove_uncommitted_export_artifacts(paths: Iterable[str]) -> None:
    seen: set[str] = set()
    for raw_path in paths:
        path = os.path.abspath(str(raw_path or "").strip())
        if not path or path in seen:
            continue
        seen.add(path)
        try:
            if os.path.isfile(path):
                os.remove(path)
        except OSError:
            continue


def _eligible_set_hash(records: list[Dict[str, Any]]) -> str:
    seed_payload = [
        {
            "record_id": str(record.get("record_id") or ""),
            "revision_id": int(record.get("revision_id") or 0),
            "revision_hash": str(record.get("revision_hash") or ""),
        }
        for record in sorted(records, key=lambda item: str(item.get("record_id") or ""))
    ]
    seed = json.dumps(seed_payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(seed.encode("utf-8")).hexdigest()


def _listing_date(payload: Dict[str, Any]) -> str:
    for key in ("挂牌开始日期", "预披露开始日期", "披露开始日期", "信息披露起始日期"):
        value = str(payload.get(key) or "").strip()
        if value:
            return value
    return ""


def ordered_export_headers(rows: Iterable[Dict[str, Any]]) -> List[str]:
    found: Dict[str, None] = {}
    for row in rows:
        for key in row.keys():
            found[str(key)] = None
    ordered = [key for key in HEADER_PRIORITY if key in found]
    ordered.extend(sorted(key for key in found if key not in HEADER_PRIORITY))
    return ordered


def _canonical_record_for_export(record: Mapping[str, Any]) -> Dict[str, Any]:
    if "canonical_record" not in record or record.get("canonical_record") is None:
        return {}
    canonical_record = record.get("canonical_record")
    if not isinstance(canonical_record, Mapping):
        raise TypeError(
            f"canonical_record must be a mapping when present, got {type(canonical_record).__name__}"
        )
    return dict(canonical_record)


def record_to_export_payload(record: Dict[str, Any]) -> Dict[str, Any]:
    """Convert a record to export payload using canonical data only.

    Export must use canonical data only - NO raw payload merge fallback.
    Missing canonical fields fail through PipelineFailure or PostProcessFinding.
    """
    canonical_record = _canonical_record_for_export(record)
    canonical_fields = canonical_record.get("canonical_fields")

    if not isinstance(canonical_fields, Mapping) or not canonical_fields:
        return {}
    payload, _ = project_canonical_record_to_export_payload(
        canonical_record,
        fail_on_missing=False,
    )
    if _record_family_of_record(record) == "deal":
        _apply_deal_date_export_semantics(payload, record)
    return payload


def _ensure_exportable_payload(
    record: Dict[str, Any],
    payload: Dict[str, Any],
    *,
    record_family: str,
) -> Dict[str, Any]:
    canonical_record = _canonical_record_for_export(record)
    canonical_fields = canonical_record.get("canonical_fields")

    if isinstance(canonical_fields, Mapping) and canonical_fields:
        project_canonical_record_to_export_payload(canonical_record, fail_on_missing=True)
        return payload

    raise ExportProjectionError(
        "record is missing canonical export data",
        failure_code="canonical_field_missing",
        missing_fields=(
            {
                "kind": "canonical",
                "field": "canonical_record",
                "canonical_field": "canonical_record",
                "export_field": "",
                "message": "canonical record data is required",
            },
        ),
    )


def _write_value_row(row: Mapping[str, Any], *, kind: str) -> List[Any]:
    if not isinstance(row, Mapping):
        raise TypeError(f"row must be a mapping, got {type(row).__name__}")
    payload = dict(row)
    field_candidates = clone_field_candidates().get(kind, {})
    values: List[Any] = []
    for header in get_output_columns_for_kind(kind):
        if header == "ID":
            continue
        matched_value = ""
        for candidate in field_candidates.get(header, [header]):
            candidate_value = payload.get(candidate)
            if not _is_export_empty_value(candidate_value):
                matched_value = candidate_value
                break
        values.append(_export_cell_value(matched_value))
    return values


def _write_workbook_default(file_path: str, rows: List[Dict[str, Any]], *, kind: str) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "records"
    headers = [header for header in get_output_columns_for_kind(kind) if header != "ID"]
    if not headers:
        headers = ordered_export_headers(rows)
    sheet.append(headers)
    for row in rows:
        if kind and kind in clone_field_candidates():
            sheet.append(_write_value_row(row, kind=kind))
        else:
            sheet.append([row.get(header, "") for header in headers])
    workbook.save(file_path)


def _deal_header_value(row: Dict[str, Any], header: str) -> Any:
    if header == "序号":
        return row.get("序号") or row.get("_row_number") or ""
    if header == "项目编号":
        return _first_non_empty_value(row, "项目编号")
    if header == "项目名称":
        return _first_non_empty_value(row, "项目名称", "标的名称")
    if header == "标的名称":
        return _first_non_empty_value(row, "标的名称", "项目名称")
    if header == "转让标的评估结果":
        return _first_non_empty_value(row, "转让标的评估结果", "转让标的评估值")
    if header == "转让标的评估值":
        return _first_non_empty_value(row, "转让标的评估值", "转让标的评估结果")
    if header == "转让底价":
        return _first_non_empty_value(row, "转让底价", "转让底价（万元）")
    if header == "转让底价（万元）":
        return _first_non_empty_value(row, "转让底价（万元）", "转让底价")
    if header == "交易价格":
        return _first_non_empty_value(row, "交易价格", "交易价格（万元）", "成交金额")
    if header == "交易价格（万元）":
        return _first_non_empty_value(row, "交易价格（万元）", "交易价格", "成交金额")
    if header == "持股占比（%）":
        return _first_non_empty_value(row, "持股占比（%）", "持股占比", "持股比例")
    if header == "持股占比":
        return _first_non_empty_value(row, "持股占比", "持股占比（%）", "持股比例")
    if header == "持股比例（%）":
        return _first_non_empty_value(row, "持股比例（%）", "持股比例", "持股占比（%）", "持股占比")
    if header == "持股比例":
        return _first_non_empty_value(row, "持股比例", "持股比例（%）", "持股占比", "持股占比（%）")
    value = row.get(header)
    return _export_cell_value(value)


def _deal_capital_sort_key(row: Dict[str, Any]) -> tuple[str, str, str, str, str]:
    return (
        str(row.get("_source_id") or ""),
        str(row.get("_group_project_code") or row.get("项目编号") or ""),
        str(row.get("_group_deal_date") or row.get("成交日期") or ""),
        f"{_sort_int(row.get('_investor_order')):08d}",
        str(row.get("投资方名称") or ""),
    )


def _merge_capital_project_cells(
    sheet,
    *,
    merge_headers: tuple[str, ...],
    headers: list[str],
    rows: list[Dict[str, Any]],
) -> None:
    if len(rows) <= 1:
        return
    merge_columns = [
        index + 1
        for index, header in enumerate(headers)
        if header in merge_headers and header not in DEAL_CAPITAL_INVESTOR_HEADERS
    ]
    if not merge_columns:
        return
    start = 0
    total = len(rows)
    while start < total:
        base = rows[start]
        base_key = (
            str(base.get("_source_id") or ""),
            str(base.get("_group_project_code") or base.get("项目编号") or ""),
            str(base.get("_group_deal_date") or base.get("成交日期") or ""),
        )
        end = start
        while end + 1 < total:
            candidate = rows[end + 1]
            candidate_key = (
                str(candidate.get("_source_id") or ""),
                str(candidate.get("_group_project_code") or candidate.get("项目编号") or ""),
                str(candidate.get("_group_deal_date") or candidate.get("成交日期") or ""),
            )
            if candidate_key != base_key:
                break
            end += 1
        if end > start:
            start_row = start + 2
            end_row = end + 2
            for column_index in merge_columns:
                sheet.merge_cells(
                    start_row=start_row,
                    end_row=end_row,
                    start_column=column_index,
                    end_column=column_index,
                )
        start = end + 1


def _write_workbook_deal(file_path: str, rows: List[Dict[str, Any]], *, output_kind: str) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    rows_by_source: Dict[str, list[Dict[str, Any]]] = {source_id: [] for source_id in DEAL_SOURCE_ORDER}
    for row in rows:
        source_id = _normalize_exchange(str(row.get("_source_id") or ""))
        if source_id not in rows_by_source:
            continue
        rows_by_source[source_id].append(dict(row))

    sheet_specs = list_deal_workbook_sheet_specs(output_kind)
    supported_sources = {spec.source_id for spec in sheet_specs}
    unsupported_sources = [
        source_id
        for source_id, source_rows in rows_by_source.items()
        if source_rows and source_id not in supported_sources
    ]
    if unsupported_sources:
        raise ExportProjectionError(
            "unsupported deal source_id for current workbook contract: "
            + ", ".join(unsupported_sources)
        )

    for spec in sheet_specs:
        source_id = spec.source_id
        headers = list(spec.headers)
        sheet = workbook.create_sheet(spec.sheet_name)
        sheet.append(headers)
        source_rows = rows_by_source[source_id]
        if output_kind == KIND_DEAL_CAPITAL:
            source_rows = sorted(source_rows, key=_deal_capital_sort_key)
        else:
            source_rows = sorted(
                source_rows,
                key=lambda row: (
                    str(row.get("_group_project_code") or row.get("项目编号") or ""),
                    str(row.get("_group_deal_date") or row.get("成交日期") or ""),
                ),
            )
        for row_number, row in enumerate(source_rows, start=1):
            row_with_number = dict(row)
            row_with_number.setdefault("_row_number", row_number)
            row_with_number.setdefault("序号", row_number)
            sheet.append([_deal_header_value(row_with_number, header) for header in headers])
        if output_kind == KIND_DEAL_CAPITAL:
            _merge_capital_project_cells(
                sheet,
                merge_headers=tuple(spec.merge_headers),
                headers=headers,
                rows=source_rows,
            )

    workbook.save(file_path)


def run_ready_export(
    store: StreamingStore,
    request: ExportRequest,
    *,
    writer=None,
    audit_action: str | None = None,
    audit_payload: Dict[str, Any] | None = None,
) -> ExportRunResult:
    from .surface_contract import SURFACE_EXPORT, scope_supported_for_surface

    record_family = _request_record_family(request)
    mode = _normalize_export_mode(request.requested_export_mode)
    output_dir = os.path.abspath(request.output_dir)
    os.makedirs(output_dir, exist_ok=True)
    cursor_id = request.cursor_id or _default_cursor_id(request, record_family=record_family)
    export_id = f"{dt.datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{uuid.uuid4().hex[:8]}"

    business_types = _normalized_business_scope_ids(
        request.business_types,
        record_family=record_family,
    )
    if not business_types:
        raise ExportProjectionError(
            f"unsupported export scope for current workbook contract: {record_family}:<empty>@{request.exchange or 'all'}"
        )
    if not scope_supported_for_surface(
        record_family=record_family,
        business_ids=business_types,
        exchange=request.exchange or "all",
        surface=SURFACE_EXPORT,
    ):
        raise ExportProjectionError(
            f"unsupported export scope for current workbook contract: {record_family}:{','.join(sorted(business_types))}@{request.exchange or 'all'}"
        )
    # Full export must not inherit incremental cursor state.
    exported = {} if mode == "full" else store.get_exported_revision_map(cursor_id)
    canonical_scope_hash = _canonical_scope_hash(request, record_family=record_family)
    export_profile_id = _export_profile_id_for_scope(
        request,
        business_types,
        record_family=record_family,
    )
    cursor_value = {} if mode == "full" else store.get_export_cursor_value(cursor_id)
    if mode == "incremental" and not cursor_value and store.has_export_history(cursor_id):
        raise ExportProjectionError(
            "incremental export cursor value is missing for existing export history",
            failure_code="incremental_cursor_value_missing",
        )
    if mode == "incremental" and cursor_value:
        expected_cursor_fields = {
            "export_profile_id": export_profile_id,
            "canonical_scope_hash": canonical_scope_hash,
            "schema_version": "schema-v1",
            "header_version": "headers-v1",
        }
        mismatched = [
            key
            for key, expected_value in expected_cursor_fields.items()
            if str(cursor_value.get(key) or "") != str(expected_value)
        ]
        if mismatched:
            raise ExportProjectionError(
                "incremental export cursor basis does not match current export scope",
                failure_code="incremental_cursor_basis_mismatch",
            )
        missing_provenance = [
            key
            for key in ("last_successful_export_id", "cursor_basis_export_id", "eligible_set_hash")
            if not str(cursor_value.get(key) or "").strip()
        ]
        if missing_provenance:
            raise ExportProjectionError(
                "incremental export cursor value is missing required basis provenance",
                failure_code="incremental_cursor_value_incomplete",
                missing_fields=[
                    {
                        "kind": "cursor_value",
                        "field": key,
                        "canonical_field": key,
                        "export_field": "",
                        "message": f"cursor value field {key} is required for incremental export",
                    }
                    for key in missing_provenance
                ],
            )
    last_successful_revision_watermark = (
        int(cursor_value.get("last_successful_revision_watermark") or 0)
        if mode == "incremental" and cursor_value
        else 0
    )
    records = _iter_records_in_export_scope(
        store,
        request,
        include_all_states=False,
        record_family=record_family,
    )
    field_missing_records = [
        record
        for record in _iter_records_in_export_scope(
            store,
            request,
            include_all_states=True,
            record_family=record_family,
        )
        if str(record.get("state") or "").strip().lower() == "field_missing"
    ]

    grouped: Dict[tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    profiles_by_group: Dict[tuple[str, str, str], Any] = {}
    record_counts_by_group: Dict[tuple[str, str, str], int] = defaultdict(int)
    marked_records: List[Dict[str, Any]] = []
    new_count = 0
    changed_count = 0

    field_missing_diagnostics: list[dict[str, Any]] = [
        _field_missing_diagnostic(record) for record in field_missing_records
    ]
    for record in records:
        record_payload_family = _record_family_of_record(record)
        if record_payload_family != record_family:
            continue
        try:
            _record_source_id(record)
        except ExportProjectionError as exc:
            field_missing_diagnostics.append(
                _field_missing_diagnostic(
                    record,
                    failure_code=getattr(exc, "failure_code", "invalid_identity_shape"),
                    missing_fields=_projection_error_missing_fields(exc),
                )
            )
            continue
        try:
            business_id = _record_business_id(record)
        except ExportProjectionError as exc:
            field_missing_diagnostics.append(
                _field_missing_diagnostic(
                    record,
                    failure_code=getattr(exc, "failure_code", "invalid_identity_shape"),
                    missing_fields=_projection_error_missing_fields(exc),
                )
            )
            continue
        profile = resolve_projection_profile(record_family, business_id) if business_id else None
        if profile is None:
            field_missing_diagnostics.append(
                _field_missing_diagnostic(
                    record,
                    failure_code="export_projection_profile_missing",
                    missing_fields=[
                        {
                            "kind": "canonical",
                            "field": "business_id",
                            "canonical_field": "business_identity.business_id",
                            "export_field": "",
                            "message": "business_id is required for export projection",
                        }
                    ],
                )
            )
            continue
        try:
            payload = _ensure_exportable_payload(
                record,
                record_to_export_payload(record),
                record_family=record_family,
            )
        except ExportProjectionError as exc:
            projection_finding = _projection_failure_finding(exc)
            if projection_finding is not None:
                _demote_ready_projection_failure(store, record, projection_finding)
            field_missing_diagnostics.append(
                _field_missing_diagnostic(
                    record,
                    failure_code=getattr(exc, "failure_code", "export_projection_failed"),
                    missing_fields=_projection_error_missing_fields(exc),
                )
            )
            continue
        if not export_evidence_verdict_accepted(resolve_artifact_evidence_verdict(record)):
            verdict = resolve_artifact_evidence_verdict(record)
            evidence_status = str(getattr(verdict, "status", "") or "").strip()
            evidence_reason_code = str(getattr(verdict, "reason_code", "") or "").strip()
            field_missing_diagnostics.append(
                _field_missing_diagnostic(
                    record,
                    failure_code="artifact_evidence_policy_rejected",
                    evidence_status=evidence_status,
                    evidence_reason_code=evidence_reason_code,
                    message=_artifact_evidence_rejection_message(evidence_status, evidence_reason_code),
                )
            )
            continue
        previous = exported.get(record["record_id"])
        bucket = "new"
        if previous is None:
            new_count += 1
        elif mode == "incremental" and int(record.get("revision_id") or 0) <= last_successful_revision_watermark:
            continue
        elif (
            int(previous.get("revision_id") or 0) != int(record.get("revision_id") or 0)
            or previous["revision_hash"] != record["revision_hash"]
        ):
            bucket = "changed"
            changed_count += 1
        else:
            continue
        marked_records.append(record)
        group_key = (profile.record_family, profile.business_id, bucket)
        profiles_by_group[group_key] = profile
        record_counts_by_group[group_key] += 1
        grouped[group_key].extend(
            _project_export_rows(
                record,
                payload,
                record_family=record_family,
                business_id=profile.business_id,
                output_kind=profile.output_kind,
            )
        )

    artifacts: List[ExportArtifact] = []
    uncommitted_artifact_paths: list[str] = []
    for (group_record_family, business_id, bucket), rows in sorted(grouped.items()):
        if not rows:
            continue
        profile = profiles_by_group[(group_record_family, business_id, bucket)]
        prefix = profile.output_stem or _safe_suffix(business_id)
        suffix = "新增" if bucket == "new" else "变更"
        file_path = os.path.join(output_dir, f"{prefix}_{suffix}_{export_id}.xlsx")
        uncommitted_artifact_paths.append(file_path)
        try:
            if writer is None:
                if group_record_family == "deal":
                    _write_workbook_deal(file_path, rows, output_kind=profile.output_kind)
                else:
                    _write_workbook_default(file_path, rows, kind=profile.output_kind)
            else:
                writer(file_path, rows)
        except Exception:
            _remove_uncommitted_export_artifacts(uncommitted_artifact_paths)
            raise
        try:
            _ensure_valid_xlsx_artifact(file_path)
        except Exception:
            _remove_uncommitted_export_artifacts(uncommitted_artifact_paths)
            raise
        artifacts.append(
            ExportArtifact(
                business_type=business_id,
                change_bucket=bucket,
                file_path=file_path,
                record_count=record_counts_by_group[(group_record_family, business_id, bucket)],
            )
        )

    try:
        revision_watermark = max((int(record.get("revision_id") or 0) for record in marked_records), default=0)
        eligible_hash = _eligible_set_hash(marked_records)
        manifest = {
            "export_id": export_id,
            "requested_export_mode": mode,
            "effective_export_mode": mode,
            "scope": _export_scope_dto(request, record_family=record_family),
            "canonical_scope_hash": canonical_scope_hash,
            "export_profile_id": export_profile_id,
            "schema_version": "schema-v1",
            "header_version": "headers-v1",
            "cursor_id": cursor_id,
            "cursor_basis": {
                "export_id": export_id,
                "eligible_set_hash": eligible_hash,
            },
            "revision_watermark": revision_watermark,
            "included_count": len(marked_records),
            "excluded_count": len(field_missing_diagnostics),
            "field_missing_blocked_records": len(field_missing_diagnostics),
            "field_missing_diagnostics": field_missing_diagnostics,
            "artifact_checksums": _artifact_checksums(artifacts),
            "records": [
                {
                    "record_id": str(record.get("record_id") or ""),
                    "revision_id": int(record.get("revision_id") or 0),
                    "revision_hash": str(record.get("revision_hash") or ""),
                }
                for record in marked_records
            ],
        }
        cursor_value = {
            "export_profile_id": export_profile_id,
            "canonical_scope_hash": canonical_scope_hash,
            "schema_version": "schema-v1",
            "header_version": "headers-v1",
            "last_successful_revision_watermark": revision_watermark,
            "last_successful_export_id": export_id,
            "cursor_basis_export_id": export_id,
            "eligible_set_hash": eligible_hash,
        }
        summary = {
            "new_records": new_count,
            "changed_records": changed_count,
            "artifacts": [artifact.file_path for artifact in artifacts],
            "requested_export_mode": mode,
            "effective_export_mode": mode,
            "cursor_id": cursor_id,
            "revision_watermark": revision_watermark,
            "field_missing_blocked_records": len(field_missing_diagnostics),
            "field_missing_diagnostics": field_missing_diagnostics,
            "manifest": manifest,
            "cursor_value": cursor_value,
        }
        if artifacts:
            resolved_audit_payload = None
            if audit_action:
                if audit_payload is None:
                    resolved_audit_payload = {}
                elif isinstance(audit_payload, Mapping):
                    resolved_audit_payload = dict(audit_payload)
                else:
                    raise TypeError("audit_payload must be a mapping when audit_action is set")
                resolved_audit_payload.update(
                    {
                        "export_id": export_id,
                        "cursor_id": cursor_id,
                        "requested_export_mode": mode,
                        "revision_watermark": revision_watermark,
                        "field_missing_blocked_records": len(field_missing_diagnostics),
                        "field_missing_diagnostics": field_missing_diagnostics,
                        "retention_count": int(request.retention_count or 20),
                        "new_records": new_count,
                        "changed_records": changed_count,
                        "artifacts": [artifact.file_path for artifact in artifacts],
                        "status": "completed",
                        "message": f"导出完成，共生成 {len(artifacts)} 个文件",
                    }
                )
            store.mark_exported(
                export_id=export_id,
                cursor_id=cursor_id,
                requested_export_mode=mode,
                date_from=request.date_from,
                date_to=request.date_to,
                project_type=",".join(sorted(business_types)),
                output_dir=output_dir,
                summary=summary,
                records=marked_records,
                manifest=manifest,
                cursor_value=cursor_value,
                audit_action=audit_action,
                audit_payload=resolved_audit_payload,
                retention_count=int(request.retention_count or 20),
            )
    except Exception:
        _remove_uncommitted_export_artifacts(uncommitted_artifact_paths)
        raise
    return ExportRunResult(
        export_id=export_id,
        cursor_id=cursor_id,
        revision_watermark=revision_watermark,
        artifacts=artifacts,
        new_records=new_count,
        changed_records=changed_count,
        field_missing_blocked_records=len(field_missing_diagnostics),
        field_missing_diagnostics=field_missing_diagnostics,
    )
